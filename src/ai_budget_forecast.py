import streamlit as st
import pandas as pd
import numpy as np
from prophet import Prophet
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings("ignore")
from datetime import datetime
import io
import traceback

# ============================================================
# PAGE CONFIG
# ============================================================
st.set_page_config(
    page_title="AI Budget Forecasting",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
* { font-family: 'Inter', sans-serif; }
.stApp { background: #0f1117; color: #e2e8f0; }
div[data-testid="stSidebar"] { background: #161b27; border-right: 1px solid #1e2d40; }
div[data-testid="stSidebar"] label { color: #94a3b8 !important; font-size: 0.82rem !important; }

.page-title { font-size: 1.5rem; font-weight: 600; color: #f1f5f9; margin-bottom: 4px; }
.page-sub   { font-size: 0.85rem; color: #64748b; margin-bottom: 28px; }
.divider    { border: none; border-top: 1px solid #1e2d40; margin: 24px 0; }

.section-label {
    font-size: 0.7rem; font-weight: 600; text-transform: uppercase;
    letter-spacing: 1px; color: #475569; margin: 24px 0 10px 0;
}
.card {
    background: #161b27; border: 1px solid #1e2d40;
    border-radius: 8px; padding: 18px; margin-bottom: 10px;
}
.card-label { font-size: 0.72rem; color: #64748b; text-transform: uppercase;
    letter-spacing: 0.5px; margin-bottom: 6px; }
.card-value { font-size: 1.3rem; font-weight: 600; color: #f1f5f9; }
.card-sub   { font-size: 0.78rem; color: #475569; margin-top: 3px; }

.col-box {
    background: #1e2d40; border: 1px solid #2d3f55;
    border-radius: 6px; padding: 12px 14px; margin-bottom: 8px;
}
.col-role   { font-size: 0.68rem; color: #3b82f6; font-weight: 600;
    text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 3px; }
.col-name   { font-size: 0.92rem; color: #e2e8f0; font-weight: 500; }
.col-sample { font-size: 0.75rem; color: #475569; margin-top: 2px; }
.col-miss   { font-size: 0.82rem; color: #ef4444; }

.info  { background: #1e2d40; border-left: 3px solid #3b82f6;
    border-radius: 4px; padding: 11px 14px; color: #94a3b8;
    font-size: 0.84rem; margin: 10px 0; }
.warn  { background: #1c1a10; border-left: 3px solid #f59e0b;
    border-radius: 4px; padding: 11px 14px; color: #fbbf24;
    font-size: 0.84rem; margin: 10px 0; }
.ok    { background: #0d1f14; border-left: 3px solid #22c55e;
    border-radius: 4px; padding: 11px 14px; color: #86efac;
    font-size: 0.84rem; margin: 10px 0; }

.fc-table { width: 100%; border-collapse: collapse; font-size: 0.84rem; margin-top: 6px; }
.fc-table th {
    background: #1e2d40; color: #94a3b8; font-weight: 500;
    padding: 9px 12px; text-align: right; font-size: 0.72rem;
    text-transform: uppercase; letter-spacing: 0.5px;
    border-bottom: 1px solid #2d3f55;
}
.fc-table th:first-child { text-align: left; }
.fc-table td { padding: 9px 12px; border-bottom: 1px solid #1a2333;
    color: #94a3b8; text-align: right; }
.fc-table td:first-child { color: #e2e8f0; font-weight: 500; text-align: left; }
.fc-table tr:hover td { background: #161b27; }
.fc-table .base { color: #f1f5f9; font-weight: 600; }
.fc-table .pos  { color: #22c55e; font-weight: 500; }
.fc-table .neg  { color: #ef4444; font-weight: 500; }
.fc-table .ev   { color: #60a5fa; font-size: 0.78rem; }

.stButton > button {
    background: #2563eb; color: white; border: none;
    border-radius: 6px; padding: 10px 20px;
    font-size: 0.86rem; font-weight: 500; width: 100%;
}
.stButton > button:hover { background: #1d4ed8; }
.stDownloadButton > button {
    background: #16a34a; color: white; border: none;
    border-radius: 6px; padding: 10px 20px;
    font-size: 0.86rem; font-weight: 500; width: 100%;
}
label { color: #94a3b8 !important; font-size: 0.82rem !important; }
</style>
""", unsafe_allow_html=True)


# ============================================================
# REGION EVENT DATES ONLY — impact learned from data
# ============================================================
REGION_EVENTS = {
    "Saudi Arabia": {
        "Ramzan":             {"month": 3,  "duration": 30},
        "Eid ul Fitr":        {"month": 4,  "duration": 4},
        "Eid ul Adha":        {"month": 6,  "duration": 4},
        "Hajj Season":        {"month": 6,  "duration": 10},
        "Saudi National Day": {"month": 9,  "duration": 2},
        "Founding Day":       {"month": 2,  "duration": 2},
        "Muharram":           {"month": 7,  "duration": 3},
    },
    "Pakistan": {
        "Ramzan":             {"month": 3,  "duration": 30},
        "Eid ul Fitr":        {"month": 4,  "duration": 4},
        "Eid ul Adha":        {"month": 6,  "duration": 4},
        "Independence Day":   {"month": 8,  "duration": 1},
        "New Year":           {"month": 1,  "duration": 2},
    },
    "UAE": {
        "Ramzan":               {"month": 3,  "duration": 30},
        "Eid ul Fitr":          {"month": 4,  "duration": 4},
        "Eid ul Adha":          {"month": 6,  "duration": 4},
        "UAE National Day":     {"month": 12, "duration": 2},
        "New Year":             {"month": 1,  "duration": 2},
        "Dubai Shopping Fest":  {"month": 1,  "duration": 30},
    },
    "Global": {
        "New Year":   {"month": 1,  "duration": 2},
        "Christmas":  {"month": 12, "duration": 7},
        "Mid-Year":   {"month": 6,  "duration": 1},
    },
}


# ============================================================
# COLUMN DETECTOR
# ============================================================
def detect_columns(df):
    det = {"date": None, "amount": None, "budget": None,
           "actual": None, "category": None, "location": None,
           "expense_type": None}

    date_kw   = ["date", "month", "period", "time", "year", "day"]
    amount_kw = ["gross_amount", "amount", "revenue", "sales", "net",
                 "value", "income", "payment", "total", "sum",
                 "amortization", "fee", "price"]
    budget_kw = ["budget", "planned", "target", "forecast"]
    actual_kw = ["actual", "real", "realized", "achieved"]
    cat_kw    = ["category", "type", "product", "family", "segment",
                 "membership", "plan", "department", "cost_center"]
    loc_kw    = ["location", "city", "branch", "store", "region", "area"]

    for col in df.columns:
        cl = col.lower().replace("_", " ")
        if not det["date"] and any(k in cl for k in date_kw):
            try:
                pd.to_datetime(df[col].dropna().iloc[0])
                det["date"] = col; continue
            except: pass
        if not det["budget"] and any(k in cl for k in budget_kw):
            if pd.api.types.is_numeric_dtype(df[col]):
                det["budget"] = col; continue
        if not det["actual"] and any(k in cl for k in actual_kw):
            if pd.api.types.is_numeric_dtype(df[col]):
                det["actual"] = col; continue
        if not det["amount"] and any(k in cl for k in amount_kw):
            if pd.api.types.is_numeric_dtype(df[col]):
                det["amount"] = col; continue
        if not det["category"] and any(k in cl for k in cat_kw):
            if df[col].dtype == object:
                det["category"] = col; continue
        if not det["location"] and any(k in cl for k in loc_kw):
            if df[col].dtype == object:
                det["location"] = col; continue

    # Check if there's a Budget/Actual type column (like "Expenses" col)
    for col in df.select_dtypes(include="object").columns:
        vals = df[col].dropna().str.strip().str.lower().unique()
        if any("budget" in v for v in vals) and any("actual" in v for v in vals):
            det["expense_type"] = col
            break

    # Fallback date
    if not det["date"]:
        for col in df.select_dtypes(include="object").columns:
            try:
                pd.to_datetime(df[col].dropna().iloc[0])
                det["date"] = col; break
            except: pass

    # Fallback amount
    if not det["amount"] and not det["budget"]:
        nums = df.select_dtypes(include="number").columns.tolist()
        if nums: det["amount"] = nums[-1]

    return det


# ============================================================
# DATA PREP — handles both formats:
# Format A: separate budget & actual columns
# Format B: one column with Budget/Actual rows (like Kaggle dataset)
# ============================================================
def prepare_data(df, det):
    """
    Returns:
        budget_df  — monthly budget totals
        actual_df  — monthly actual totals  
        combined   — for forecasting
        data_type  — 'budget_actual' or 'timeseries'
    """
    df = df.copy()

    # Format B: Budget/Actual in rows (expense_type column exists)
    if det["expense_type"]:
        date_col = det["date"]
        exp_col  = det["expense_type"]
        df[date_col] = pd.to_datetime(df[date_col])

        # Numeric columns = cost categories
        num_cols = df.select_dtypes(include="number").columns.tolist()

        budget_rows = df[df[exp_col].str.lower().str.strip() == "budget"].copy()
        actual_rows = df[df[exp_col].str.lower().str.strip() == "actual"].copy()

        budget_rows["_total"] = budget_rows[num_cols].sum(axis=1)
        actual_rows["_total"] = actual_rows[num_cols].sum(axis=1)

        budget_df = budget_rows[[date_col, "_total"]].rename(
            columns={date_col: "Month", "_total": "Budget"})
        actual_df = actual_rows[[date_col, "_total"]].rename(
            columns={date_col: "Month", "_total": "Actual"})

        # Category breakdown
        cat_budget = budget_rows[num_cols].mean()
        cat_actual = actual_rows[num_cols].mean()

        return budget_df, actual_df, actual_rows, num_cols, "budget_actual"

    # Format A: time series (one amount column)
    amount_col = det["amount"] or det["budget"]
    date_col   = det["date"]
    df[date_col] = pd.to_datetime(df[date_col])
    df["_month"] = df[date_col].dt.to_period("M")
    agg = df.groupby("_month")[amount_col].sum().reset_index()
    agg.columns = ["Month", "Actual"]
    agg["Month"] = agg["Month"].dt.to_timestamp()

    return None, agg, df, [amount_col], "timeseries"


# ============================================================
# FORECAST ENGINE
# AI decides method based on data size
# ============================================================
def run_forecast(actual_df, date_col, amount_col,
                 forecast_months, region_events, custom_events):
    """
    < 24 months  → Trend + Seasonal average (more reliable)
    >= 24 months → Prophet (learns patterns automatically)
    Impact of events is ALWAYS learned from data, never hard-coded.
    """
    ts = actual_df.copy()
    ts.columns = ["ds", "y"]
    ts["ds"] = pd.to_datetime(ts["ds"])
    ts = ts[ts["y"] >= 0].sort_values("ds").reset_index(drop=True)

    n_months = len(ts)
    method_used = ""

    if n_months >= 24:
        # Prophet — learns seasonality & event impact from data
        method_used = "Prophet (AI — learned from data)"

        years = list(range(ts["ds"].dt.year.min(), datetime.now().year + 3))
        holiday_rows = []
        for name, ev in region_events.items():
            for yr in years:
                try:
                    start = pd.Timestamp(f"{yr}-{ev['month']:02d}-01")
                    for d in pd.date_range(start=start,
                                           periods=ev["duration"], freq="D"):
                        holiday_rows.append({
                            "holiday": name, "ds": d,
                            "lower_window": -1, "upper_window": 2
                        })
                except: pass
        for ev in custom_events:
            try:
                for d in pd.date_range(start=ev["date"],
                                       periods=ev["duration"], freq="D"):
                    holiday_rows.append({
                        "holiday": ev["name"], "ds": d,
                        "lower_window": -1, "upper_window": 2
                    })
            except: pass

        hdf = pd.DataFrame(holiday_rows) if holiday_rows else None

        model = Prophet(
            holidays=hdf,
            yearly_seasonality=True,
            weekly_seasonality=False,
            daily_seasonality=False,
            seasonality_mode="multiplicative",
            changepoint_prior_scale=0.1,
        )
        model.fit(ts)

        future   = model.make_future_dataframe(
            periods=forecast_months, freq="MS")
        forecast = model.predict(future)

        last = ts["ds"].max()
        ff   = forecast[forecast["ds"] > last][
            ["ds", "yhat", "yhat_lower", "yhat_upper"]].copy()
        ff   = ff.head(forecast_months).reset_index(drop=True)

        # Learn event impact from model components
        learned_impact = {}
        try:
            comp_cols = [c for c in forecast.columns
                         if c in (hdf["holiday"].unique() if hdf is not None else [])]
            for ev_name in comp_cols:
                non_zero = forecast[forecast[ev_name] != 0][ev_name]
                if len(non_zero) > 0:
                    learned_impact[ev_name] = round(float(non_zero.mean()) * 100, 1)
        except: pass

        result = pd.DataFrame({
            "Month":     ff["ds"],
            "Low":       ff["yhat_lower"].clip(lower=0).round(0),
            "Forecast":  ff["yhat"].clip(lower=0).round(0),
            "High":      ff["yhat_upper"].clip(lower=0).round(0),
        })

    else:
        # Trend + Seasonal method — reliable for small datasets
        method_used = f"Trend + Seasonal Average (data: {n_months} months)"
        learned_impact = {}

        avg   = ts["y"].mean()
        std   = ts["y"].std()
        trend = 0
        if n_months >= 3:
            trend = (ts["y"].iloc[-3:].mean() - ts["y"].iloc[:3].mean()) / n_months

        # Monthly seasonal index from historical data
        ts["month_num"] = ts["ds"].dt.month
        seasonal = ts.groupby("month_num")["y"].mean()
        global_avg = ts["y"].mean()
        seasonal_idx = (seasonal / global_avg).fillna(1.0)

        rows = []
        last = ts["ds"].max()
        for i in range(forecast_months):
            next_month = last + pd.DateOffset(months=i + 1)
            m_num      = next_month.month
            s_idx      = seasonal_idx.get(m_num, 1.0)
            base       = (avg + trend * (n_months + i)) * s_idx
            base       = max(base, 0)
            rows.append({
                "Month":    next_month,
                "Low":      max(0, base - std * s_idx),
                "Forecast": base,
                "High":     base + std * s_idx,
            })

        result = pd.DataFrame(rows)

    # Month number for event mapping
    result["month_num"] = pd.to_datetime(result["Month"]).dt.month
    result["Month_str"] = pd.to_datetime(result["Month"]).dt.strftime("%Y-%m")

    return result, method_used, learned_impact


# ============================================================
# EXCEL BUILDER — 2 clean sheets only
# ============================================================
def build_excel(fc_df, budget_df, actual_df, ts,
                det, region, currency, method_used,
                learned_impact, region_events, custom_events):

    wb  = Workbook()
    wb.remove(wb.active)

    # Colors
    BG    = "0f1923"
    HDR   = "1e2d40"
    ROW1  = "111827"
    ROW2  = "0f1117"
    GBG   = "0d2318"; GFG = "22c55e"
    RBG   = "1f0d0d"; RFG = "ef4444"
    BBG   = "0d1f38"; BFG = "60a5fa"
    WH    = "f1f5f9"; MU  = "475569"
    AC    = "1e3a5f"

    def C(ws, ref, val, bg=None, fg=WH, bold=False,
          fmt=None, al="left", sz=10):
        c = ws[ref]; c.value = val
        if bg: c.fill = PatternFill("solid", fgColor=bg)
        c.font = Font(bold=bold, color=fg, size=sz, name="Calibri")
        c.alignment = Alignment(horizontal=al, vertical="center", wrap_text=True)
        if fmt: c.number_format = fmt

    def BR(ws, rng, col="1e2d40"):
        s = Side(style="thin", color=col)
        for row in ws[rng]:
            for c in row:
                c.border = Border(left=s, right=s, top=s, bottom=s)

    def CW(ws, c, w): ws.column_dimensions[get_column_letter(c)].width = w
    def RH(ws, r, h): ws.row_dimensions[r].height = h

    # Month → event map
    ev_map = {}
    for name, ev in region_events.items():
        ev_map[ev["month"]] = name
    for ev in custom_events:
        try: ev_map[pd.Timestamp(ev["date"]).month] = ev["name"]
        except: pass

    # ── SHEET 1: Budget Forecast ────────────────────────────
    ws1 = wb.create_sheet("Budget Forecast")
    ws1.sheet_view.showGridLines = False

    ws1.merge_cells("A1:H1")
    C(ws1, "A1", "AI Budget Forecasting Framework",
      bg=BG, fg=WH, bold=True, sz=13)
    RH(ws1, 1, 34)

    ws1.merge_cells("A2:H2")
    C(ws1, "A2",
      f"Generated: {datetime.now().strftime('%d %b %Y')}   |   "
      f"Region: {region}   |   Currency: {currency}   |   "
      f"Method: {method_used}",
      bg=BG, fg=MU, sz=9)
    RH(ws1, 2, 18)

    # KPIs
    total_fc   = fc_df["Forecast"].sum()
    avg_fc     = fc_df["Forecast"].mean()
    peak_month = fc_df.loc[fc_df["Forecast"].idxmax(), "Month_str"]

    kpi_data = [
        ("Total Forecast",   f"{currency} {total_fc:,.0f}", "A3", "B3", "A4", "B4"),
        ("Monthly Average",  f"{currency} {avg_fc:,.0f}",   "C3", "D3", "C4", "D4"),
        ("Peak Month",       str(peak_month),                "E3", "F3", "E4", "F4"),
        ("Forecast Period",  f"{len(fc_df)} months",        "G3", "H3", "G4", "H4"),
    ]
    for label, val, ls, le, vs, ve in kpi_data:
        ws1.merge_cells(f"{ls}:{le}")
        ws1.merge_cells(f"{vs}:{ve}")
        C(ws1, ls, label, bg=AC, fg=MU, sz=8, al="center")
        C(ws1, vs, val,   bg=HDR, fg=WH, bold=True, sz=11, al="center")
        RH(ws1, 3, 18); RH(ws1, 4, 26)

    RH(ws1, 5, 6)

    # Headers
    hdrs = ["Month", "Low Case", "Base Forecast", "High Case",
            "Event", "AI-Learned Impact", "MoM Change", "Status"]
    for ci, h in enumerate(hdrs, 1):
        C(ws1, f"{get_column_letter(ci)}6", h,
          bg=HDR, fg=MU, bold=True, sz=9, al="center")
    RH(ws1, 6, 20)
    BR(ws1, "A6:H6")

    prev = None
    for ri, row in fc_df.iterrows():
        r  = 7 + ri
        bg = ROW1 if ri % 2 == 0 else ROW2
        base = row["Forecast"]
        low  = row["Low"]
        high = row["High"]
        mom  = ((base - prev) / prev * 100) if prev else 0
        prev = base
        mn   = row["month_num"]
        ev   = ev_map.get(mn, "")
        imp  = learned_impact.get(ev)
        imp_str = f"{imp:+.1f}% learned" if imp and ev else ""
        status = "Above Avg" if base > avg_fc * 1.1 else (
                 "Below Avg" if base < avg_fc * 0.9 else "On Track")
        mbg = GBG if mom >= 0 else RBG
        mfg = GFG if mom >= 0 else RFG

        C(ws1, f"A{r}", row["Month_str"], bg=bg, fg=WH, bold=True, al="center")
        C(ws1, f"B{r}", low,  bg=bg, fg=MU, fmt="#,##0", al="right")
        C(ws1, f"C{r}", base, bg=bg, fg=WH, bold=True, fmt="#,##0", al="right")
        C(ws1, f"D{r}", high, bg=bg, fg=MU, fmt="#,##0", al="right")
        C(ws1, f"E{r}", ev,   bg=BBG if ev else bg, fg=BFG if ev else MU,
          al="center", sz=9)
        C(ws1, f"F{r}", imp_str, bg=BBG if imp_str else bg,
          fg=BFG if imp_str else MU, al="center", sz=9)
        C(ws1, f"G{r}", f"{mom:+.1f}%", bg=mbg, fg=mfg, bold=True,
          al="center", sz=9)
        C(ws1, f"H{r}", status, bg=bg, fg=MU, al="center", sz=9)
        RH(ws1, r, 19)

    BR(ws1, f"A6:H{6 + len(fc_df)}")

    # Scenario footer
    fr = 8 + len(fc_df)
    RH(ws1, fr - 1, 8)
    for label, val, bg, fg, sc, ec in [
        ("Worst Case (Low)",  fc_df["Low"].sum(),      RBG, RFG, "A", "B"),
        ("Base Forecast",     fc_df["Forecast"].sum(), BBG, BFG, "C", "E"),
        ("Best Case (High)",  fc_df["High"].sum(),     GBG, GFG, "F", "H"),
    ]:
        ws1.merge_cells(f"{sc}{fr}:{ec}{fr}")
        ws1.merge_cells(f"{sc}{fr+1}:{ec}{fr+1}")
        C(ws1, f"{sc}{fr}",   label, bg=bg, fg=fg, sz=8, al="center")
        C(ws1, f"{sc}{fr+1}", f"{currency} {val:,.0f}",
          bg=bg, fg=fg, bold=True, sz=11, al="center")
        RH(ws1, fr, 18); RH(ws1, fr + 1, 26)

    for c, w in [(1,12),(2,15),(3,18),(4,15),(5,18),(6,20),(7,12),(8,12)]:
        CW(ws1, c, w)

    # ── SHEET 2: Variance Analysis ──────────────────────────
    ws2 = wb.create_sheet("Variance Analysis")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:F1")
    C(ws2, "A1", "Variance Analysis — Budget vs Actual",
      bg=BG, fg=WH, bold=True, sz=12)
    RH(ws2, 1, 30)

    ws2.merge_cells("A2:F2")
    C(ws2, "A2",
      "Historical actual performance compared to budget. "
      "Variance is calculated from your data — not estimated.",
      bg=BG, fg=MU, sz=9)
    RH(ws2, 2, 16)
    RH(ws2, 3, 6)

    if budget_df is not None and actual_df is not None:
        merged = pd.merge(
            budget_df.rename(columns={"Budget": "Budget"}),
            actual_df.rename(columns={"Actual": "Actual"}),
            on="Month", how="inner"
        )
        merged["Variance"]     = merged["Actual"] - merged["Budget"]
        merged["Variance_Pct"] = (merged["Variance"] / merged["Budget"] * 100).round(1)
        merged["Month_str"]    = pd.to_datetime(merged["Month"]).dt.strftime("%Y-%m")

        for ci, h in enumerate(
            ["Month", "Budget", "Actual", "Variance",
             "Variance %", "Status"], 1):
            C(ws2, f"{get_column_letter(ci)}4", h,
              bg=HDR, fg=MU, bold=True, sz=9, al="center")
        RH(ws2, 4, 20)

        for ri, row in merged.iterrows():
            r   = 5 + ri
            bg  = ROW1 if ri % 2 == 0 else ROW2
            var = row["Variance"]
            vp  = row["Variance_Pct"]
            vbg = GBG if var >= 0 else RBG
            vfg = GFG if var >= 0 else RFG
            st  = "Above Budget" if var >= 0 else "Below Budget"

            C(ws2, f"A{r}", row["Month_str"], bg=bg,  fg=WH, bold=True, al="center")
            C(ws2, f"B{r}", row["Budget"],    bg=bg,  fg=MU, fmt="#,##0", al="right")
            C(ws2, f"C{r}", row["Actual"],    bg=bg,  fg=WH, bold=True, fmt="#,##0", al="right")
            C(ws2, f"D{r}", var,              bg=vbg, fg=vfg, fmt="#,##0", al="right", bold=True)
            C(ws2, f"E{r}", f"{vp:+.1f}%",   bg=vbg, fg=vfg, al="center", bold=True)
            C(ws2, f"F{r}", st,               bg=vbg, fg=vfg, al="center")
            RH(ws2, r, 19)

        BR(ws2, f"A4:F{4 + len(merged)}")

        # Summary
        sr = 6 + len(merged)
        ws2.merge_cells(f"A{sr}:F{sr}")
        C(ws2, f"A{sr}", "SUMMARY", bg=HDR, fg=MU, bold=True, sz=9, al="center")
        RH(ws2, sr, 18)

        for ri2, (label, val, bg, fg) in enumerate([
            ("Total Budget",   merged["Budget"].sum(),   ROW1, MU),
            ("Total Actual",   merged["Actual"].sum(),   ROW1, WH),
            ("Total Variance", merged["Variance"].sum(),
             GBG if merged["Variance"].sum() >= 0 else RBG,
             GFG if merged["Variance"].sum() >= 0 else RFG),
        ]):
            r3 = sr + 1 + ri2
            C(ws2, f"A{r3}", label, bg=bg, fg=fg, al="left")
            ws2.merge_cells(f"B{r3}:F{r3}")
            C(ws2, f"B{r3}", val, bg=bg, fg=fg, bold=True,
              fmt="#,##0", al="right")
            RH(ws2, r3, 20)

    else:
        # Time series — show last 12 months actual trend
        ts_copy = ts.copy()
        ts_copy[det["date"]] = pd.to_datetime(ts_copy[det["date"]])
        ts_copy["_m"] = ts_copy[det["date"]].dt.to_period("M")
        am_col = det["amount"] or det["budget"]
        am = ts_copy.groupby("_m")[am_col].sum().reset_index()
        am.columns = ["Month", "Actual"]
        am = am.tail(12).reset_index(drop=True)
        avg_b = am["Actual"].mean()

        for ci, h in enumerate(
            ["Month", "Actual", "Avg Benchmark",
             "Variance", "Variance %", "Status"], 1):
            C(ws2, f"{get_column_letter(ci)}4", h,
              bg=HDR, fg=MU, bold=True, sz=9, al="center")
        RH(ws2, 4, 20)

        for ri, row in am.iterrows():
            r   = 5 + ri
            bg  = ROW1 if ri % 2 == 0 else ROW2
            var = row["Actual"] - avg_b
            vp  = (var / avg_b * 100) if avg_b else 0
            vbg = GBG if var >= 0 else RBG
            vfg = GFG if var >= 0 else RFG
            st  = "Above Avg" if var >= 0 else "Below Avg"

            C(ws2, f"A{r}", str(row["Month"]), bg=bg,  fg=WH, bold=True, al="center")
            C(ws2, f"B{r}", row["Actual"],     bg=bg,  fg=WH, bold=True, fmt="#,##0", al="right")
            C(ws2, f"C{r}", avg_b,             bg=bg,  fg=MU, fmt="#,##0", al="right")
            C(ws2, f"D{r}", var,               bg=vbg, fg=vfg, fmt="#,##0", al="right", bold=True)
            C(ws2, f"E{r}", f"{vp:+.1f}%",    bg=vbg, fg=vfg, al="center", bold=True)
            C(ws2, f"F{r}", st,                bg=vbg, fg=vfg, al="center")
            RH(ws2, r, 19)

        BR(ws2, f"A4:F{4 + len(am)}")

    for c, w in [(1,12),(2,18),(3,18),(4,16),(5,13),(6,15)]:
        CW(ws2, c, w)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ============================================================
# MAIN APP
# ============================================================
def main():
    st.markdown('<div class="page-title">AI Budget Forecasting</div>', unsafe_allow_html=True)
    st.markdown('<div class="page-sub">Upload any revenue or budget CSV — AI learns patterns and generates rolling forecast automatically</div>', unsafe_allow_html=True)

    # ── SIDEBAR ─────────────────────────────────────────────
    with st.sidebar:
        st.markdown("**Settings**")
        st.markdown("---")

        region = st.selectbox("Region",
            ["Saudi Arabia", "Pakistan", "UAE", "Global"])
        currency = st.selectbox("Currency",
            ["SAR", "PKR", "AED", "USD", "EUR", "GBP"])
        forecast_years  = st.slider("Forecast period (years)", 1, 10, 1)
        forecast_months = forecast_years * 12

        st.markdown("---")
        st.markdown("**Events**")
        st.markdown(
            "<div style='font-size:0.78rem;color:#475569;margin-bottom:8px'>"
            "Select which events to consider. Impact is learned from your data automatically."
            "</div>", unsafe_allow_html=True)

        region_evs = REGION_EVENTS.get(region, {})
        selected_events = {}
        for ev_name, ev_data in region_evs.items():
            if st.checkbox(ev_name, value=True, key=f"ev_{ev_name}"):
                selected_events[ev_name] = ev_data

        st.markdown("---")
        st.markdown("**Custom Events**")
        if "custom_events" not in st.session_state:
            st.session_state.custom_events = []

        with st.expander("Add custom event"):
            cn  = st.text_input("Event name", placeholder="e.g. Company Anniversary")
            cd  = st.date_input("Start date")
            cdu = st.number_input("Duration (days)", 1, 90, 1)
            if st.button("Add"):
                if cn:
                    st.session_state.custom_events.append({
                        "name": cn, "date": str(cd), "duration": int(cdu)})

        for ev in st.session_state.custom_events:
            st.markdown(
                f"<div style='font-size:0.78rem;color:#64748b'>"
                f"{ev['name']} — {ev['date']}</div>",
                unsafe_allow_html=True)

        if st.session_state.custom_events:
            if st.button("Clear"):
                st.session_state.custom_events = []

    # ── STEP 1: UPLOAD ───────────────────────────────────────
    st.markdown('<div class="section-label">Upload Data</div>', unsafe_allow_html=True)

    uploaded = st.file_uploader(
        "Upload CSV or Excel file", type=["csv", "xlsx", "xls"],
        label_visibility="collapsed")

    if not uploaded:
        st.markdown('<div class="info">Upload any CSV or Excel file with revenue, sales, or budget data. AI will detect columns and learn patterns automatically.</div>',
                    unsafe_allow_html=True)
        return

    try:
        if uploaded.name.endswith(".csv"):
            df = pd.read_csv(uploaded)
        else:
            df = pd.read_excel(uploaded)
    except Exception as e:
        st.error(f"Could not read file: {e}")
        return

    st.markdown(f'<div class="ok">File loaded — {len(df):,} rows, {len(df.columns)} columns</div>',
                unsafe_allow_html=True)

    # ── STEP 2: COLUMN DETECTION ─────────────────────────────
    st.markdown('<div class="section-label">Column Detection</div>', unsafe_allow_html=True)

    det = detect_columns(df)

    c1, c2, c3, c4 = st.columns(4)
    for cont, role, key in [
        (c1, "Date",     "date"),
        (c2, "Amount",   "amount"),
        (c3, "Budget",   "budget"),
        (c4, "Type Col", "expense_type"),
    ]:
        with cont:
            col = det.get(key)
            if col:
                sample = str(df[col].dropna().iloc[0])[:20]
                st.markdown(f"""<div class="col-box">
                    <div class="col-role">{role}</div>
                    <div class="col-name">{col}</div>
                    <div class="col-sample">e.g. {sample}</div>
                </div>""", unsafe_allow_html=True)
            else:
                st.markdown(f"""<div class="col-box">
                    <div class="col-role">{role}</div>
                    <div class="col-miss">Not detected</div>
                </div>""", unsafe_allow_html=True)

    with st.expander("Override column selection"):
        all_cols = list(df.columns)
        det["date"] = st.selectbox("Date column", all_cols,
            index=all_cols.index(det["date"]) if det["date"] in all_cols else 0)
        lo = ["None"] + all_cols
        am = st.selectbox("Amount column", lo,
            index=all_cols.index(det["amount"]) + 1 if det["amount"] in all_cols else 0)
        det["amount"] = None if am == "None" else am

    # ── STEP 3: GENERATE ─────────────────────────────────────
    st.markdown('<div class="section-label">Generate Forecast</div>', unsafe_allow_html=True)

    c1, c2, c3 = st.columns(3)
    with c1:
        st.markdown(f"""<div class="card">
            <div class="card-label">Region</div>
            <div class="card-value">{region}</div>
        </div>""", unsafe_allow_html=True)
    with c2:
        st.markdown(f"""<div class="card">
            <div class="card-label">Forecast Period</div>
            <div class="card-value">{forecast_years} Year(s)</div>
            <div class="card-sub">{forecast_months} months rolling</div>
        </div>""", unsafe_allow_html=True)
    with c3:
        ev_count = len(selected_events) + len(st.session_state.custom_events)
        st.markdown(f"""<div class="card">
            <div class="card-label">Events</div>
            <div class="card-value">{ev_count}</div>
            <div class="card-sub">Impact learned from data</div>
        </div>""", unsafe_allow_html=True)

    with st.expander("Preview data"):
        st.dataframe(df.head(8), use_container_width=True)

    if st.button("Generate Forecast"):
        with st.spinner("AI is analyzing your data..."):
            try:
                budget_df, actual_df, raw, num_cols, dtype = prepare_data(df, det)

                # Build time series for forecasting
                if dtype == "budget_actual":
                    ts_df = actual_df.rename(columns={"Month": "ds", "Actual": "y"})
                else:
                    ts_df = actual_df.rename(columns={"Month": "ds", "Actual": "y"})

                fc_df, method, learned_impact = run_forecast(
                    ts_df, "ds", "y",
                    forecast_months,
                    selected_events,
                    st.session_state.custom_events
                )

                # Results
                st.markdown(f'<div class="ok">Forecast complete — Method: {method}</div>',
                            unsafe_allow_html=True)

                m1, m2, m3, m4 = st.columns(4)
                for cont, label, val, sub in [
                    (m1, "Total Forecast",   f"{currency} {fc_df['Forecast'].sum():,.0f}", f"{forecast_months}M total"),
                    (m2, "Monthly Average",  f"{currency} {fc_df['Forecast'].mean():,.0f}", "base case"),
                    (m3, "High Case Total",  f"{currency} {fc_df['High'].sum():,.0f}", "best scenario"),
                    (m4, "Low Case Total",   f"{currency} {fc_df['Low'].sum():,.0f}", "worst scenario"),
                ]:
                    with cont:
                        st.markdown(f"""<div class="card">
                            <div class="card-label">{label}</div>
                            <div class="card-value">{val}</div>
                            <div class="card-sub">{sub}</div>
                        </div>""", unsafe_allow_html=True)

                # Forecast table
                st.markdown('<div class="section-label">Forecast Preview</div>', unsafe_allow_html=True)

                ev_map = {}
                for name, ev in selected_events.items():
                    ev_map[ev["month"]] = name
                for ev in st.session_state.custom_events:
                    try: ev_map[pd.Timestamp(ev["date"]).month] = ev["name"]
                    except: pass

                avg_fc = fc_df["Forecast"].mean()
                rows   = ""
                prev   = None
                for _, row in fc_df.iterrows():
                    base = row["Forecast"]
                    mom  = ((base - prev) / prev * 100) if prev else 0
                    prev = base
                    ev   = ev_map.get(row["month_num"], "")
                    imp  = learned_impact.get(ev)
                    imp_str = f"{imp:+.1f}%" if imp and ev else ""
                    mc   = "pos" if mom >= 0 else "neg"
                    rows += f"""<tr>
                        <td>{row['Month_str']}</td>
                        <td>{currency} {row['Low']:,.0f}</td>
                        <td class="base">{currency} {base:,.0f}</td>
                        <td>{currency} {row['High']:,.0f}</td>
                        <td class="ev">{ev}</td>
                        <td class="ev">{imp_str}</td>
                        <td class="{mc}">{mom:+.1f}%</td>
                    </tr>"""

                st.markdown(f"""
                <table class="fc-table">
                    <thead><tr>
                        <th>Month</th>
                        <th>Low Case</th>
                        <th>Base Forecast</th>
                        <th>High Case</th>
                        <th>Event</th>
                        <th>Learned Impact</th>
                        <th>MoM Change</th>
                    </tr></thead>
                    <tbody>{rows}</tbody>
                </table>""", unsafe_allow_html=True)

                st.markdown("<br>", unsafe_allow_html=True)

                excel = build_excel(
                    fc_df, budget_df, actual_df, df,
                    det, region, currency, method,
                    learned_impact, selected_events,
                    st.session_state.custom_events
                )
                fname = f"Budget_Forecast_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
                st.download_button(
                    "Download Excel Report",
                    data=excel, file_name=fname,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

            except Exception as e:
                st.error(f"Error: {e}")
                st.code(traceback.format_exc())


if __name__ == "__main__":
    main()