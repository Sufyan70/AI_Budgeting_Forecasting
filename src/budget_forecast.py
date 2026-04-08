import pandas as pd
import numpy as np
from prophet import Prophet
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')
from datetime import datetime
import os

def get_saudi_events(years):
    events = []
    for year in years:
        events += [
            {'holiday': 'Ramzan',           'ds': pd.date_range(start=f'{year}-03-01', periods=30, freq='D')},
            {'holiday': 'Eid_ul_Fitr',      'ds': pd.date_range(start=f'{year}-04-01', periods=4,  freq='D')},
            {'holiday': 'Eid_ul_Adha',      'ds': pd.date_range(start=f'{year}-06-16', periods=4,  freq='D')},
            {'holiday': 'Hajj_Season',      'ds': pd.date_range(start=f'{year}-06-10', periods=10, freq='D')},
            {'holiday': 'Saudi_National_Day','ds': pd.date_range(start=f'{year}-09-23', periods=2, freq='D')},
            {'holiday': 'Muharram',         'ds': pd.date_range(start=f'{year}-07-07', periods=3,  freq='D')},
            {'holiday': 'Founding_Day',     'ds': pd.date_range(start=f'{year}-02-22', periods=2,  freq='D')},
        ]
    rows = []
    for e in events:
        for d in e['ds']:
            rows.append({'holiday': e['holiday'], 'ds': d, 'lower_window': -1, 'upper_window': 2})
    return pd.DataFrame(rows)



DARK_GREEN   = "1A5276"
MID_GREEN    = "1E8449"
LIGHT_GREEN  = "D5F5E3"
ORANGE       = "E67E22"
LIGHT_ORANGE = "FDEBD0"
WHITE        = "FFFFFF"
GRAY         = "F2F3F4"
RED          = "C0392B"
LIGHT_RED    = "FADBD8"
BLUE         = "2E86C1"
LIGHT_BLUE   = "D6EAF8"

def hdr(ws, cell, val, bg=DARK_GREEN, fg=WHITE, bold=True, size=11, align='center'):
    c = ws[cell]
    c.value = val
    c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(bold=bold, color=fg, size=size)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)

def cell_val(ws, cell, val, bg=None, bold=False, fmt=None, align='center', color="000000"):
    c = ws[cell]
    c.value = val
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(bold=bold, color=color)
    c.alignment = Alignment(horizontal=align, vertical='center')
    if fmt:
        c.number_format = fmt

def add_border(ws, cell_range):
    thin = Side(style='thin')
    for row in ws[cell_range]:
        for c in row:
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def row_height(ws, row, height):
    ws.row_dimensions[row].height = height



def build_summary_sheet(wb, monthly_fc, total_actual):
    ws = wb.create_sheet("Executive Summary")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    hdr(ws, "A1", "MAWGIF — AI BUDGET FORECASTING REPORT", DARK_GREEN, WHITE, True, 14)
    row_height(ws, 1, 35)

    ws.merge_cells("A2:G2")
    hdr(ws, "A2", f"Generated: {datetime.now().strftime('%d %B %Y')}  |  Forecast: Next 12 Months  |  Saudi Events Adjusted",
        MID_GREEN, WHITE, False, 10)
    row_height(ws, 2, 20)

    ws.merge_cells("A4:G4")
    hdr(ws, "A4", "KEY PERFORMANCE INDICATORS", DARK_GREEN, WHITE, True, 11)
    row_height(ws, 4, 25)

    kpis = [
        ("Total Historical Revenue", f"SAR {total_actual:,.0f}", LIGHT_BLUE),
        ("12M Base Forecast",        f"SAR {monthly_fc['Forecast'].sum():,.0f}", LIGHT_GREEN),
        ("12M High Case",            f"SAR {monthly_fc['High_Case'].sum():,.0f}", LIGHT_ORANGE),
        ("Avg Monthly (Base)",       f"SAR {monthly_fc['Forecast'].mean():,.0f}", GRAY),
        ("Peak Forecast Month",      str(monthly_fc.loc[monthly_fc['Forecast'].idxmax(), 'Month']), LIGHT_ORANGE),
        ("Lowest Forecast Month",    str(monthly_fc.loc[monthly_fc['Forecast'].idxmin(), 'Month']), LIGHT_RED),
    ]

    positions = [(5,1),(5,3),(5,5),(7,1),(7,3),(7,5)]
    for i, ((r, col_s), (label, val, bg)) in enumerate(zip(positions, kpis)):
        col_e = col_s + 1
        ws.merge_cells(f"{get_column_letter(col_s)}{r}:{get_column_letter(col_e)}{r}")
        ws.merge_cells(f"{get_column_letter(col_s)}{r+1}:{get_column_letter(col_e)}{r+1}")
        hdr(ws, f"{get_column_letter(col_s)}{r}", label, MID_GREEN, WHITE, False, 9)
        cell_val(ws, f"{get_column_letter(col_s)}{r+1}", val, bg, True, align='center')
        row_height(ws, r, 20)
        row_height(ws, r+1, 22)

    ws.merge_cells("A10:G10")
    hdr(ws, "A10", "MONTHLY FORECAST SUMMARY", DARK_GREEN, WHITE, True, 11)
    row_height(ws, 10, 25)

    for ci, h in enumerate(["Month", "Low Case (SAR)", "Base Forecast (SAR)",
                             "High Case (SAR)", "MoM Change %", "Saudi Event", "Status"], 1):
        hdr(ws, f"{get_column_letter(ci)}11", h, MID_GREEN, WHITE)
        row_height(ws, 11, 22)

    saudi_events_months = {
        '2026-02': 'Founding Day',
        '2026-03': 'Ramzan',
        '2026-04': 'Eid ul Fitr',
        '2026-06': 'Hajj + Eid Adha',
        '2026-07': 'Muharram',
        '2026-09': 'National Day',
    }

    prev_fc = None
    for ri, row_data in monthly_fc.head(12).iterrows():
        r = 12 + ri
        mom = ((row_data['Forecast'] - prev_fc) / prev_fc * 100) if prev_fc else 0
        prev_fc = row_data['Forecast']
        bg = LIGHT_GREEN if ri % 2 == 0 else WHITE
        event = saudi_events_months.get(str(row_data['Month']), '-')
        status = " Peak" if row_data['Forecast'] > monthly_fc['Forecast'].mean() * 1.3 else (
                 " Low"  if row_data['Forecast'] < monthly_fc['Forecast'].mean() * 0.7 else " Normal")

        cell_val(ws, f"A{r}", str(row_data['Month']),   bg, align='center')
        cell_val(ws, f"B{r}", row_data['Low_Case'],     bg, fmt='#,##0', align='right')
        cell_val(ws, f"C{r}", row_data['Forecast'],     bg, fmt='#,##0', bold=True, align='right')
        cell_val(ws, f"D{r}", row_data['High_Case'],    bg, fmt='#,##0', align='right')
        cell_val(ws, f"E{r}", f"{mom:+.1f}%",
                 LIGHT_GREEN if mom >= 0 else LIGHT_RED, align='center')
        cell_val(ws, f"F{r}", event, LIGHT_ORANGE if event != '-' else bg, align='center')
        cell_val(ws, f"G{r}", status, bg, align='center')
        row_height(ws, r, 20)

    add_border(ws, f"A11:G{11+len(monthly_fc.head(12))}")
    for col, w in [(1,14),(2,18),(3,20),(4,18),(5,14),(6,18),(7,12)]:
        col_width(ws, col, w)



def build_monthly_sheet(wb, monthly_fc):
    ws = wb.create_sheet("Monthly Forecast")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    hdr(ws, "A1", "MONTHLY REVENUE FORECAST — BASE / LOW / HIGH SCENARIOS", DARK_GREEN, WHITE, True, 12)
    row_height(ws, 1, 30)

    for ci, h in enumerate(["Month", "Low Case", "Base Forecast", "High Case",
                             "Low vs Base", "High vs Base", "Recommended Budget"], 1):
        hdr(ws, f"{get_column_letter(ci)}2", h, MID_GREEN, WHITE)
        row_height(ws, 2, 22)

    for ri, row_data in monthly_fc.head(12).iterrows():
        r = 3 + ri
        bg = LIGHT_GREEN if ri % 2 == 0 else WHITE
        low_diff  = row_data['Low_Case'] - row_data['Forecast']
        high_diff = row_data['High_Case'] - row_data['Forecast']
        recommended = (row_data['Forecast'] + row_data['High_Case']) / 2

        cell_val(ws, f"A{r}", str(row_data['Month']), bg, align='center')
        cell_val(ws, f"B{r}", row_data['Low_Case'],   bg, fmt='#,##0', align='right')
        cell_val(ws, f"C{r}", row_data['Forecast'],   bg, fmt='#,##0', bold=True, align='right')
        cell_val(ws, f"D{r}", row_data['High_Case'],  bg, fmt='#,##0', align='right')
        cell_val(ws, f"E{r}", f"{low_diff:+,.0f}",   LIGHT_RED,    align='right')
        cell_val(ws, f"F{r}", f"{high_diff:+,.0f}",  LIGHT_GREEN,  align='right')
        cell_val(ws, f"G{r}", recommended,            LIGHT_ORANGE, fmt='#,##0', bold=True, align='right')
        row_height(ws, r, 20)

    add_border(ws, f"A2:G{2+len(monthly_fc.head(12))}")
    for col, w in [(1,14),(2,18),(3,18),(4,18),(5,16),(6,16),(7,20)]:
        col_width(ws, col, w)



def build_location_sheet(wb, daily_gross, monthly_fc):
    ws = wb.create_sheet("Location Forecast")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    hdr(ws, "A1", "LOCATION-WISE FORECAST — AHP | JAP | DCP | JWM", DARK_GREEN, WHITE, True, 12)
    row_height(ws, 1, 30)

    paid = daily_gross[daily_gross['payment_status'] == 'PAID']
    loc_shares = paid.groupby('location')['gross_amount'].sum()
    total = loc_shares.sum()
    loc_shares = loc_shares / total

    colors = {'AHP': LIGHT_BLUE, 'JAP': LIGHT_GREEN, 'DCP': LIGHT_ORANGE, 'JWM': LIGHT_RED}

    for ci, h in enumerate(["Location", "Historical Share %", "Jan 2026",
                             "Feb 2026", "Mar 2026", "Q1 Total", "Q2 Total", "Full Year"], 1):
        hdr(ws, f"{get_column_letter(ci)}2", h, MID_GREEN, WHITE)

    for ri, (loc, share) in enumerate(loc_shares.items()):
        r = 3 + ri
        bg = colors.get(loc, WHITE)

        def get_fc(month_str):
            v = monthly_fc[monthly_fc['Month_str'] == month_str]['Forecast'].values
            return float(v[0]) * share if len(v) else 0

        jan = get_fc('2026-01')
        feb = get_fc('2026-02')
        mar = get_fc('2026-03')
        q1  = monthly_fc.head(3)['Forecast'].sum() * share
        q2  = monthly_fc.iloc[3:6]['Forecast'].sum() * share
        full = monthly_fc.head(12)['Forecast'].sum() * share

        cell_val(ws, f"A{r}", loc,              bg, bold=True, align='center')
        cell_val(ws, f"B{r}", f"{share*100:.1f}%", bg, align='center')
        cell_val(ws, f"C{r}", jan,              bg, fmt='#,##0', align='right')
        cell_val(ws, f"D{r}", feb,              bg, fmt='#,##0', align='right')
        cell_val(ws, f"E{r}", mar,              bg, fmt='#,##0', align='right')
        cell_val(ws, f"F{r}", q1,               bg, fmt='#,##0', bold=True, align='right')
        cell_val(ws, f"G{r}", q2,               bg, fmt='#,##0', bold=True, align='right')
        cell_val(ws, f"H{r}", full,             bg, fmt='#,##0', bold=True, align='right')
        row_height(ws, r, 22)

    add_border(ws, f"A2:H{2+len(loc_shares)}")
    for col, w in [(1,12),(2,18),(3,15),(4,15),(5,15),(6,16),(7,16),(8,20)]:
        col_width(ws, col, w)



def build_variance_sheet(wb, daily_gross):
    ws = wb.create_sheet("Budget vs Actual")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    hdr(ws, "A1", "BUDGET vs ACTUAL — VARIANCE ANALYSIS (Historical)", DARK_GREEN, WHITE, True, 12)
    row_height(ws, 1, 30)

    paid = daily_gross[daily_gross['payment_status'] == 'PAID'].copy()
    paid['month'] = paid['date'].dt.to_period('M')
    actual_monthly = paid.groupby('month')['gross_amount'].sum().reset_index()
    actual_monthly.columns = ['Month', 'Actual']
    actual_monthly = actual_monthly.tail(12).reset_index(drop=True)
    avg_budget = actual_monthly['Actual'].mean()

    for ci, h in enumerate(["Month", "Actual Revenue", "Budget (Avg)",
                             "Variance (SAR)", "Variance %", "Status"], 1):
        hdr(ws, f"{get_column_letter(ci)}2", h, MID_GREEN, WHITE)

    for ri, row_data in actual_monthly.iterrows():
        r = 3 + ri
        variance = row_data['Actual'] - avg_budget
        var_pct  = (variance / avg_budget * 100) if avg_budget else 0
        status   = "✅ Above Budget" if variance >= 0 else "⚠️ Below Budget"
        bg_var   = LIGHT_GREEN if variance >= 0 else LIGHT_RED
        bg       = GRAY if ri % 2 == 0 else WHITE

        cell_val(ws, f"A{r}", str(row_data['Month']),  bg,     align='center')
        cell_val(ws, f"B{r}", row_data['Actual'],      bg,     fmt='#,##0', align='right')
        cell_val(ws, f"C{r}", avg_budget,              bg,     fmt='#,##0', align='right')
        cell_val(ws, f"D{r}", variance,                bg_var, fmt='#,##0', align='right')
        cell_val(ws, f"E{r}", f"{var_pct:+.1f}%",     bg_var, align='center')
        cell_val(ws, f"F{r}", status,                  bg_var, align='center')
        row_height(ws, r, 20)

    add_border(ws, f"A2:F{2+len(actual_monthly)}")
    for col, w in [(1,14),(2,18),(3,16),(4,18),(5,14),(6,18)]:
        col_width(ws, col, w)



def build_events_sheet(wb):
    ws = wb.create_sheet("Saudi Events Impact")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    hdr(ws, "A1", "SAUDI EVENTS — FORECAST IMPACT CALENDAR 2026", DARK_GREEN, WHITE, True, 12)
    row_height(ws, 1, 30)

    for ci, h in enumerate(["Event", "Approximate Date 2026",
                             "Duration", "Expected Revenue Impact", "CFO Recommendation"], 1):
        hdr(ws, f"{get_column_letter(ci)}2", h, MID_GREEN, WHITE)

    events = [
        ("Founding Day",       "22 Feb 2026",   "2 days",   "+10% to +15%",  "Increase budget allocation"),
        ("Ramzan Start",       "~18 Mar 2026",  "30 days",  "+20% to +35%",  "Peak season — plan max capacity"),
        ("Eid ul Fitr",        "~17 Apr 2026",  "4 days",   "+30% to +50%",  "Highest revenue — all hands on deck"),
        ("Hajj Season",        "~10 Jun 2026",  "10 days",  "+15% to +25%",  "Corporate memberships spike"),
        ("Eid ul Adha",        "~16 Jun 2026",  "4 days",   "+25% to +40%",  "Second peak — prepare in advance"),
        ("Muharram",           "~7 Jul 2026",   "3 days",   "-5% to +5%",    "Neutral — monitor closely"),
        ("Saudi National Day", "23 Sep 2026",   "2 days",   "+10% to +20%",  "Promotional campaign opportunity"),
    ]

    event_colors = [LIGHT_GREEN, LIGHT_ORANGE, LIGHT_ORANGE,
                    LIGHT_BLUE, LIGHT_ORANGE, GRAY, LIGHT_BLUE]

    for ri, (ev, dt, dur, impact, rec) in enumerate(events):
        r = 3 + ri
        bg = event_colors[ri]
        cell_val(ws, f"A{r}", ev,     bg, bold=True, align='left')
        cell_val(ws, f"B{r}", dt,     bg, align='center')
        cell_val(ws, f"C{r}", dur,    bg, align='center')
        cell_val(ws, f"D{r}", impact, bg, align='center')
        cell_val(ws, f"E{r}", rec,    bg, align='left')
        row_height(ws, r, 22)

    add_border(ws, f"A2:E{2+len(events)}")
    for col, w in [(1,22),(2,22),(3,14),(4,24),(5,35)]:
        col_width(ws, col, w)



def build_scenarios_sheet(wb, monthly_fc):
    ws = wb.create_sheet("Scenario Planning")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:D1")
    hdr(ws, "A1", "SCENARIO PLANNING — BEST / BASE / WORST CASE", DARK_GREEN, WHITE, True, 12)
    row_height(ws, 1, 30)

    scenarios = [
        ("BEST CASE",  "All events drive peak revenue, full capacity, no disruptions",
         monthly_fc['High_Case'].sum(),  LIGHT_GREEN),
        ("BASE CASE",  "Normal operations with Saudi seasonal adjustments",
         monthly_fc['Forecast'].sum(),   LIGHT_BLUE),
        ("WORST CASE", "Economic slowdown, low event impact, membership drop",
         monthly_fc['Low_Case'].sum(),   LIGHT_RED),
    ]

    for ci, h in enumerate(["Scenario", "Assumptions", "12M Total (SAR)", "Avg Monthly (SAR)"], 1):
        hdr(ws, f"{get_column_letter(ci)}2", h, MID_GREEN, WHITE)

    for ri, (name, assumption, total, bg) in enumerate(scenarios):
        r = 3 + ri
        cell_val(ws, f"A{r}", name,        bg, bold=True, align='center')
        cell_val(ws, f"B{r}", assumption,  bg, align='left')
        cell_val(ws, f"C{r}", total,       bg, fmt='#,##0', bold=True, align='right')
        cell_val(ws, f"D{r}", total/12,    bg, fmt='#,##0', align='right')
        row_height(ws, r, 25)

    ws.merge_cells("A7:D7")
    hdr(ws, "A7", "MONTHLY SCENARIO BREAKDOWN", MID_GREEN, WHITE, True, 11)
    row_height(ws, 7, 25)

    for ci, h in enumerate(["Month", "Worst Case (SAR)", "Base Case (SAR)", "Best Case (SAR)"], 1):
        hdr(ws, f"{get_column_letter(ci)}8", h, DARK_GREEN, WHITE)

    for ri, row_data in monthly_fc.head(12).iterrows():
        r = 9 + ri
        bg = GRAY if ri % 2 == 0 else WHITE
        cell_val(ws, f"A{r}", str(row_data['Month']), bg,          align='center')
        cell_val(ws, f"B{r}", row_data['Low_Case'],   LIGHT_RED,   fmt='#,##0', align='right')
        cell_val(ws, f"C{r}", row_data['Forecast'],   LIGHT_BLUE,  fmt='#,##0', bold=True, align='right')
        cell_val(ws, f"D{r}", row_data['High_Case'],  LIGHT_GREEN, fmt='#,##0', align='right')
        row_height(ws, r, 20)

    add_border(ws, f"A8:D{8+len(monthly_fc.head(12))}")
    for col, w in [(1,14),(2,20),(3,20),(4,20)]:
        col_width(ws, col, w)



def run_forecast(data_folder="data", output_folder="outputs"):
    print("=" * 55)
    print("  MAWGIF AI BUDGET FORECASTING TOOL")
    print("=" * 55)

    print("\n Loading company data...")
    daily_gross = pd.read_csv(f"{data_folder}/daily_gross_amount_and_members.csv")
    daily_amort = pd.read_csv(f"{data_folder}/daily_amortization_tm1.csv")
    monthly     = pd.read_csv(f"{data_folder}/monthly_amortization_dimensional.csv")

    daily_gross['date'] = pd.to_datetime(daily_gross['date'])
    daily_amort['date'] = pd.to_datetime(daily_amort['date'])
    monthly['year_month'] = pd.to_datetime(monthly['year_month'])
    print(" Data loaded successfully!")

    print("\n Training AI model with Saudi events...")
    paid = daily_gross[daily_gross['payment_status'] == 'PAID'].copy()
    prophet_df = paid.groupby('date')['gross_amount'].sum().reset_index()
    prophet_df.columns = ['ds', 'y']
    prophet_df = prophet_df[prophet_df['ds'] <= '2026-01-18']

    saudi_holidays = get_saudi_events(list(range(2021, 2028)))

    model = Prophet(
        holidays=saudi_holidays,
        yearly_seasonality=True,
        weekly_seasonality=True,
        daily_seasonality=False,
        seasonality_mode='multiplicative',
        holidays_prior_scale=15,
        changepoint_prior_scale=0.1
    )
    model.fit(prophet_df)

    future   = model.make_future_dataframe(periods=365)
    forecast = model.predict(future)
    print(" AI model trained & forecast ready!")

    print("\n Building monthly forecast tables...")
    fc_future = forecast[forecast['ds'] > '2026-01-18'][
        ['ds', 'yhat', 'yhat_lower', 'yhat_upper']].copy()
    fc_future['month'] = fc_future['ds'].dt.to_period('M')
    monthly_fc = fc_future.groupby('month').agg(
        {'yhat': 'sum', 'yhat_lower': 'sum', 'yhat_upper': 'sum'}).reset_index()
    monthly_fc.columns = ['Month', 'Forecast', 'Low_Case', 'High_Case']
    monthly_fc['Forecast']  = monthly_fc['Forecast'].clip(lower=0)
    monthly_fc['Low_Case']  = monthly_fc['Low_Case'].clip(lower=0)
    monthly_fc['High_Case'] = monthly_fc['High_Case'].clip(lower=0)
    monthly_fc['Month_str'] = monthly_fc['Month'].astype(str)

    total_actual = paid['gross_amount'].sum()

    print("\n Building Excel report...")
    wb = Workbook()
    wb.remove(wb.active)

    build_summary_sheet(wb, monthly_fc, total_actual)
    build_monthly_sheet(wb, monthly_fc)
    build_location_sheet(wb, daily_gross, monthly_fc)
    build_variance_sheet(wb, daily_gross)
    build_events_sheet(wb)
    build_scenarios_sheet(wb, monthly_fc)

    os.makedirs(output_folder, exist_ok=True)
    fname = f"{output_folder}/Mawgif_Budget_Forecast_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    wb.save(fname)

    print(f"\n{'=' * 55}")
    print(f"   EXCEL REPORT READY!")
    print(f"   Saved: {fname}")
    print(f"{'=' * 55}")
    print(f"\n  12-Month Forecast Summary:")
    print(f"  Base Case :  SAR {monthly_fc.head(12)['Forecast'].sum():>15,.0f}")
    print(f"  High Case :  SAR {monthly_fc.head(12)['High_Case'].sum():>15,.0f}")
    print(f"  Low Case  :  SAR {monthly_fc.head(12)['Low_Case'].sum():>15,.0f}")
    print(f"{'=' * 55}\n")


if __name__ == "__main__":
    run_forecast()