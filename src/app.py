# import streamlit as st
# import pandas as pd
# import numpy as np
# from prophet import Prophet
# from openpyxl import Workbook
# from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
# from openpyxl.utils import get_column_letter
# import warnings
# warnings.filterwarnings('ignore')
# from datetime import datetime
# import io
# import traceback


# st.set_page_config(
#     page_title="Mawgif AI Budget Forecasting",
#     page_icon="📊",
#     layout="wide"
# )

# st.markdown("""
# <style>
# @import url('https://fonts.googleapis.com/css2?family=Syne:wght@400;600;700;800&family=DM+Sans:wght@300;400;500&display=swap');

# * { font-family: 'DM Sans', sans-serif; }
# h1, h2, h3 { font-family: 'Syne', sans-serif; }

# .stApp { background: #0a0f1e; color: #e8eaf0; }

# .hero {
#     background: linear-gradient(135deg, #0d1b35 0%, #1a2f5e 50%, #0d1b35 100%);
#     border: 1px solid rgba(99,179,237,0.2);
#     border-radius: 20px;
#     padding: 48px;
#     margin-bottom: 32px;
#     text-align: center;
#     position: relative;
#     overflow: hidden;
# }
# .hero::before {
#     content: '';
#     position: absolute;
#     top: -50%;
#     left: -50%;
#     width: 200%;
#     height: 200%;
#     background: radial-gradient(circle at center, rgba(99,179,237,0.05) 0%, transparent 60%);
#     pointer-events: none;
# }
# .hero h1 {
#     font-size: 2.8rem;
#     font-weight: 800;
#     color: #ffffff;
#     margin: 0 0 8px 0;
#     letter-spacing: -1px;
# }
# .hero p {
#     color: #90a4c8;
#     font-size: 1.1rem;
#     margin: 0;
#     font-weight: 300;
# }
# .hero .accent { color: #63b3ed; }

# .step-card {
#     background: #111827;
#     border: 1px solid rgba(99,179,237,0.15);
#     border-radius: 14px;
#     padding: 24px;
#     margin-bottom: 16px;
#     transition: border-color 0.2s;
# }
# .step-card:hover { border-color: rgba(99,179,237,0.4); }
# .step-num {
#     display: inline-block;
#     background: #63b3ed;
#     color: #0a0f1e;
#     font-family: 'Syne', sans-serif;
#     font-weight: 800;
#     width: 32px;
#     height: 32px;
#     border-radius: 50%;
#     text-align: center;
#     line-height: 32px;
#     margin-bottom: 12px;
#     font-size: 0.9rem;
# }
# .step-card h3 { color: #ffffff; margin: 0 0 6px 0; font-size: 1rem; font-weight: 600; }
# .step-card p  { color: #6b7fa3; margin: 0; font-size: 0.85rem; }

# .col-card {
#     background: #111827;
#     border: 1px solid rgba(99,179,237,0.15);
#     border-radius: 12px;
#     padding: 20px;
#     margin-bottom: 12px;
# }
# .col-tag {
#     display: inline-block;
#     background: rgba(99,179,237,0.15);
#     color: #63b3ed;
#     padding: 3px 10px;
#     border-radius: 20px;
#     font-size: 0.78rem;
#     font-weight: 600;
#     margin-bottom: 8px;
# }
# .col-name { color: #ffffff; font-weight: 600; font-size: 1rem; margin-bottom: 4px; }
# .col-sample { color: #6b7fa3; font-size: 0.82rem; }

# .metric-box {
#     background: linear-gradient(135deg, #111827, #1a2535);
#     border: 1px solid rgba(99,179,237,0.2);
#     border-radius: 14px;
#     padding: 24px;
#     text-align: center;
# }
# .metric-label { color: #6b7fa3; font-size: 0.82rem; font-weight: 500; margin-bottom: 8px; text-transform: uppercase; letter-spacing: 0.5px; }
# .metric-value { color: #63b3ed; font-family: 'Syne', sans-serif; font-weight: 800; font-size: 1.8rem; }
# .metric-sub   { color: #4a5f80; font-size: 0.78rem; margin-top: 4px; }

# .success-banner {
#     background: linear-gradient(135deg, #064e3b, #065f46);
#     border: 1px solid #10b981;
#     border-radius: 14px;
#     padding: 24px 32px;
#     text-align: center;
#     margin: 24px 0;
# }
# .success-banner h2 { color: #10b981; font-family: 'Syne', sans-serif; margin: 0 0 8px 0; }
# .success-banner p  { color: #6ee7b7; margin: 0; }

# .warning-box {
#     background: rgba(245,158,11,0.1);
#     border: 1px solid rgba(245,158,11,0.3);
#     border-radius: 10px;
#     padding: 16px;
#     color: #fbbf24;
#     font-size: 0.88rem;
#     margin: 12px 0;
# }
# .info-box {
#     background: rgba(99,179,237,0.08);
#     border: 1px solid rgba(99,179,237,0.2);
#     border-radius: 10px;
#     padding: 16px;
#     color: #90c4e8;
#     font-size: 0.88rem;
#     margin: 12px 0;
# }

# div[data-testid="stFileUploader"] {
#     background: #111827;
#     border: 2px dashed rgba(99,179,237,0.3);
#     border-radius: 14px;
#     padding: 8px;
# }
# div[data-testid="stFileUploader"]:hover {
#     border-color: rgba(99,179,237,0.6);
# }

# .stButton > button {
#     background: linear-gradient(135deg, #2563eb, #1d4ed8);
#     color: white;
#     border: none;
#     border-radius: 10px;
#     padding: 14px 32px;
#     font-family: 'Syne', sans-serif;
#     font-weight: 700;
#     font-size: 1rem;
#     width: 100%;
#     cursor: pointer;
#     transition: all 0.2s;
# }
# .stButton > button:hover {
#     background: linear-gradient(135deg, #3b82f6, #2563eb);
#     transform: translateY(-1px);
# }

# .stSelectbox > div { background: #111827; border-color: rgba(99,179,237,0.2); }
# label { color: #90a4c8 !important; }

# hr { border-color: rgba(99,179,237,0.1); }
# </style>
# """, unsafe_allow_html=True)



# def get_saudi_events(years):
#     rows = []
#     for year in years:
#         events = [
#             ('Ramzan',            f'{year}-03-01', 30, -1, 3),
#             ('Eid_ul_Fitr',       f'{year}-04-01',  4, -2, 3),
#             ('Eid_ul_Adha',       f'{year}-06-16',  4, -2, 3),
#             ('Hajj_Season',       f'{year}-06-10', 10, -1, 2),
#             ('Saudi_National_Day',f'{year}-09-23',  2, -1, 2),
#             ('Muharram',          f'{year}-07-07',  3, -1, 1),
#             ('Founding_Day',      f'{year}-02-22',  2, -1, 2),
#         ]
#         for name, start, periods, lw, uw in events:
#             for d in pd.date_range(start=start, periods=periods, freq='D'):
#                 rows.append({'holiday': name, 'ds': d,
#                              'lower_window': lw, 'upper_window': uw})
#     return pd.DataFrame(rows)



# def detect_columns(df):
#     """AI-style column detection — works on ANY CSV"""
#     detected = {'date': None, 'amount': None, 'category': None,
#                  'location': None, 'members': None}

#     date_keywords    = ['date','dt','time','period','month','year','day','week']
#     amount_keywords  = ['amount','revenue','sales','gross','net','value','income',
#                         'amortization','payment','fee','price','total','sum']
#     category_keywords= ['category','type','product','family','segment','class',
#                         'membership','plan','tier','group']
#     location_keywords= ['location','city','branch','store','region','area',
#                         'site','place','zone','country']
#     member_keywords  = ['member','count','customer','client','user','subscriber',
#                         'active','headcount']

#     for col in df.columns:
#         col_lower = col.lower().replace('_',' ').replace('-',' ')

#         if detected['date'] is None:
#             if any(k in col_lower for k in date_keywords):
#                 try:
#                     pd.to_datetime(df[col].dropna().iloc[0])
#                     detected['date'] = col
#                     continue
#                 except: pass

#         if detected['amount'] is None:
#             if any(k in col_lower for k in amount_keywords):
#                 if pd.api.types.is_numeric_dtype(df[col]):
#                     detected['amount'] = col
#                     continue

#         if detected['category'] is None:
#             if any(k in col_lower for k in category_keywords):
#                 if df[col].dtype == object:
#                     detected['category'] = col
#                     continue

#         if detected['location'] is None:
#             if any(k in col_lower for k in location_keywords):
#                 if df[col].dtype == object:
#                     detected['location'] = col
#                     continue

#         if detected['members'] is None:
#             if any(k in col_lower for k in member_keywords):
#                 if pd.api.types.is_numeric_dtype(df[col]):
#                     detected['members'] = col

#     if detected['date'] is None:
#         for col in df.select_dtypes(include='object').columns:
#             try:
#                 pd.to_datetime(df[col].dropna().iloc[0])
#                 detected['date'] = col
#                 break
#             except: pass

#     if detected['amount'] is None:
#         num_cols = df.select_dtypes(include='number').columns.tolist()
#         if num_cols:
#             detected['amount'] = num_cols[0]

#     return detected



# DARK   = "1A3A5C"
# MID    = "1E6B8A"
# LGREEN = "D5F5E3"
# LBLUE  = "D6EAF8"
# LORG   = "FDEBD0"
# LRED   = "FADBD8"
# GRAY   = "F2F3F4"
# WHITE  = "FFFFFF"

# def _h(ws, cell, val, bg=None, fg="FFFFFF", bold=True, size=10, align='center'):
#     c = ws[cell]
#     c.value = val
#     if bg: c.fill = PatternFill("solid", fgColor=bg)
#     c.font = Font(bold=bold, color=fg, size=size)
#     c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)

# def _v(ws, cell, val, bg=None, bold=False, fmt=None, align='center', fg="000000"):
#     c = ws[cell]
#     c.value = val
#     if bg: c.fill = PatternFill("solid", fgColor=bg)
#     c.font = Font(bold=bold, color=fg)
#     c.alignment = Alignment(horizontal=align, vertical='center')
#     if fmt: c.number_format = fmt

# def _border(ws, rng):
#     s = Side(style='thin')
#     for row in ws[rng]:
#         for c in row:
#             c.border = Border(left=s, right=s, top=s, bottom=s)

# def _cw(ws, col, w): ws.column_dimensions[get_column_letter(col)].width = w
# def _rh(ws, row, h): ws.row_dimensions[row].height = h


# def build_excel(monthly_fc, df_clean, detected, forecast_periods=12):
#     wb = Workbook()
#     wb.remove(wb.active)

#     total_actual = df_clean[detected['amount']].sum()
#     saudi_map = {
#         '02': 'Founding Day', '03': 'Ramzan', '04': 'Eid ul Fitr',
#         '06': 'Hajj + Eid Adha', '07': 'Muharram', '09': 'National Day'
#     }

#     ws = wb.create_sheet("Executive Summary")
#     ws.sheet_view.showGridLines = False

#     ws.merge_cells("A1:H1")
#     _h(ws, "A1", "MAWGIF — AI BUDGET FORECASTING REPORT", DARK, "FFFFFF", True, 14)
#     _rh(ws, 1, 38)

#     ws.merge_cells("A2:H2")
#     _h(ws, "A2",
#        f"Generated: {datetime.now().strftime('%d %B %Y')}  |  "
#        f"Forecast: Next {forecast_periods} Months  |  Saudi Events Adjusted",
#        MID, "FFFFFF", False, 10)
#     _rh(ws, 2, 22)

#     # KPIs
#     ws.merge_cells("A4:H4")
#     _h(ws, "A4", "KEY PERFORMANCE INDICATORS", DARK, "FFFFFF", True, 11)
#     _rh(ws, 4, 26)

#     fc12 = monthly_fc.head(forecast_periods)
#     kpis = [
#         ("Total Historical Revenue", f"SAR {total_actual:,.0f}", LBLUE),
#         ("12M Base Forecast",        f"SAR {fc12['Forecast'].sum():,.0f}", LGREEN),
#         ("12M High Case",            f"SAR {fc12['High_Case'].sum():,.0f}", LORG),
#         ("Avg Monthly (Base)",       f"SAR {fc12['Forecast'].mean():,.0f}", GRAY),
#         ("Peak Month",   str(fc12.loc[fc12['Forecast'].idxmax(), 'Month']), LORG),
#         ("Lowest Month", str(fc12.loc[fc12['Forecast'].idxmin(), 'Month']), LRED),
#     ]
#     positions = [(5,1),(5,3),(5,5),(7,1),(7,3),(7,5)]
#     for (r, cs), (label, val, bg) in zip(positions, kpis):
#         ce = cs + 1
#         ws.merge_cells(f"{get_column_letter(cs)}{r}:{get_column_letter(ce)}{r}")
#         ws.merge_cells(f"{get_column_letter(cs)}{r+1}:{get_column_letter(ce)}{r+1}")
#         _h(ws, f"{get_column_letter(cs)}{r}", label, MID, "FFFFFF", False, 9)
#         _v(ws, f"{get_column_letter(cs)}{r+1}", val, bg, True, align='center')
#         _rh(ws, r, 20); _rh(ws, r+1, 24)

#     ws.merge_cells("A10:H10")
#     _h(ws, "A10", "MONTHLY FORECAST OVERVIEW", DARK, "FFFFFF", True, 11)
#     _rh(ws, 10, 26)

#     for ci, h in enumerate(["Month","Low Case (SAR)","Base Forecast (SAR)",
#                              "High Case (SAR)","MoM Change","Saudi Event","Status"], 1):
#         _h(ws, f"{get_column_letter(ci)}11", h, MID, "FFFFFF")
#         _rh(ws, 11, 22)

#     prev = None
#     for ri, row in fc12.iterrows():
#         r = 12 + ri
#         mom = ((row['Forecast'] - prev) / prev * 100) if prev else 0
#         prev = row['Forecast']
#         bg = LGREEN if ri % 2 == 0 else WHITE
#         month_str = str(row['Month'])
#         mm = month_str[-2:] if len(month_str) >= 7 else ''
#         event = saudi_map.get(mm, '—')
#         mean_fc = fc12['Forecast'].mean()
#         status = (" Peak"   if row['Forecast'] > mean_fc * 1.3 else
#                   " Low"    if row['Forecast'] < mean_fc * 0.7 else
#                   " Normal")
#         _v(ws, f"A{r}", month_str,         bg, align='center')
#         _v(ws, f"B{r}", row['Low_Case'],   bg, fmt='#,##0', align='right')
#         _v(ws, f"C{r}", row['Forecast'],   bg, fmt='#,##0', bold=True, align='right')
#         _v(ws, f"D{r}", row['High_Case'],  bg, fmt='#,##0', align='right')
#         _v(ws, f"E{r}", f"{mom:+.1f}%",
#            LGREEN if mom >= 0 else LRED, align='center')
#         _v(ws, f"F{r}", event, LORG if event != '—' else bg, align='center')
#         _v(ws, f"G{r}", status, bg, align='center')
#         _rh(ws, r, 20)

#     _border(ws, f"A11:G{11+len(fc12)}")
#     for col, w in [(1,14),(2,18),(3,20),(4,18),(5,12),(6,18),(7,13)]:
#         _cw(ws, col, w)

#     ws2 = wb.create_sheet("Monthly Forecast")
#     ws2.sheet_view.showGridLines = False
#     ws2.merge_cells("A1:G1")
#     _h(ws2, "A1", "MONTHLY REVENUE FORECAST — LOW / BASE / HIGH SCENARIOS",
#        DARK, "FFFFFF", True, 12)
#     _rh(ws2, 1, 32)

#     for ci, h in enumerate(["Month","Low Case","Base Forecast","High Case",
#                              "Low vs Base","High vs Base","Recommended Budget"], 1):
#         _h(ws2, f"{get_column_letter(ci)}2", h, MID, "FFFFFF")
#         _rh(ws2, 2, 22)

#     for ri, row in fc12.iterrows():
#         r = 3 + ri
#         bg = LGREEN if ri % 2 == 0 else WHITE
#         recommended = (row['Forecast'] + row['High_Case']) / 2
#         _v(ws2, f"A{r}", str(row['Month']),                bg, align='center')
#         _v(ws2, f"B{r}", row['Low_Case'],                  bg, fmt='#,##0', align='right')
#         _v(ws2, f"C{r}", row['Forecast'],                  bg, fmt='#,##0', bold=True, align='right')
#         _v(ws2, f"D{r}", row['High_Case'],                 bg, fmt='#,##0', align='right')
#         _v(ws2, f"E{r}", row['Low_Case'] - row['Forecast'],LRED,  fmt='#,##0', align='right')
#         _v(ws2, f"F{r}", row['High_Case']- row['Forecast'],LGREEN,fmt='#,##0', align='right')
#         _v(ws2, f"G{r}", recommended, LORG, fmt='#,##0', bold=True, align='right')
#         _rh(ws2, r, 20)

#     _border(ws2, f"A2:G{2+len(fc12)}")
#     for col, w in [(1,14),(2,16),(3,18),(4,16),(5,16),(6,16),(7,20)]:
#         _cw(ws2, col, w)

#     if detected['location']:
#         ws3 = wb.create_sheet("Location Forecast")
#         ws3.sheet_view.showGridLines = False
#         ws3.merge_cells("A1:H1")
#         _h(ws3, "A1", f"LOCATION-WISE FORECAST — by {detected['location'].upper()}",
#            DARK, "FFFFFF", True, 12)
#         _rh(ws3, 1, 32)

#         loc_col = detected['location']
#         amt_col = detected['amount']
#         loc_shares = df_clean.groupby(loc_col)[amt_col].sum()
#         total_s = loc_shares.sum()
#         loc_shares = loc_shares / total_s

#         colors_list = [LBLUE, LGREEN, LORG, LRED, GRAY, LGREEN, LBLUE]
#         for ci, h in enumerate(["Location","Historical Share %","Month 1",
#                                  "Month 2","Month 3","Q1 Total","Q2 Total","Full Year"], 1):
#             _h(ws3, f"{get_column_letter(ci)}2", h, MID, "FFFFFF")

#         for ri, (loc, share) in enumerate(loc_shares.items()):
#             r = 3 + ri
#             bg = colors_list[ri % len(colors_list)]
#             m1 = fc12.iloc[0]['Forecast'] * share if len(fc12) > 0 else 0
#             m2 = fc12.iloc[1]['Forecast'] * share if len(fc12) > 1 else 0
#             m3 = fc12.iloc[2]['Forecast'] * share if len(fc12) > 2 else 0
#             q1 = fc12.head(3)['Forecast'].sum() * share
#             q2 = fc12.iloc[3:6]['Forecast'].sum() * share
#             full = fc12['Forecast'].sum() * share
#             _v(ws3, f"A{r}", str(loc),          bg, bold=True, align='center')
#             _v(ws3, f"B{r}", f"{share*100:.1f}%",bg, align='center')
#             _v(ws3, f"C{r}", m1, bg, fmt='#,##0', align='right')
#             _v(ws3, f"D{r}", m2, bg, fmt='#,##0', align='right')
#             _v(ws3, f"E{r}", m3, bg, fmt='#,##0', align='right')
#             _v(ws3, f"F{r}", q1, bg, fmt='#,##0', bold=True, align='right')
#             _v(ws3, f"G{r}", q2, bg, fmt='#,##0', bold=True, align='right')
#             _v(ws3, f"H{r}", full, bg, fmt='#,##0', bold=True, align='right')
#             _rh(ws3, r, 22)

#         _border(ws3, f"A2:H{2+len(loc_shares)}")
#         for col, w in [(1,16),(2,18),(3,15),(4,15),(5,15),(6,16),(7,16),(8,20)]:
#             _cw(ws3, col, w)

#     ws4 = wb.create_sheet("Budget vs Actual")
#     ws4.sheet_view.showGridLines = False
#     ws4.merge_cells("A1:F1")
#     _h(ws4, "A1", "BUDGET vs ACTUAL — VARIANCE ANALYSIS", DARK, "FFFFFF", True, 12)
#     _rh(ws4, 1, 32)

#     date_col = detected['date']
#     amt_col  = detected['amount']
#     df_hist  = df_clean.copy()
#     df_hist[date_col] = pd.to_datetime(df_hist[date_col])
#     df_hist['_month'] = df_hist[date_col].dt.to_period('M')
#     actual_m = df_hist.groupby('_month')[amt_col].sum().reset_index()
#     actual_m.columns = ['Month', 'Actual']
#     actual_m = actual_m.tail(12).reset_index(drop=True)
#     avg_budget = actual_m['Actual'].mean()

#     for ci, h in enumerate(["Month","Actual Revenue","Budget Target (Avg)",
#                              "Variance (SAR)","Variance %","Status"], 1):
#         _h(ws4, f"{get_column_letter(ci)}2", h, MID, "FFFFFF")

#     for ri, row in actual_m.iterrows():
#         r = 3 + ri
#         var = row['Actual'] - avg_budget
#         var_pct = (var / avg_budget * 100) if avg_budget else 0
#         status  = " Above Budget" if var >= 0 else "⚠️ Below Budget"
#         bg_var  = LGREEN if var >= 0 else LRED
#         bg      = GRAY if ri % 2 == 0 else WHITE
#         _v(ws4, f"A{r}", str(row['Month']), bg,     align='center')
#         _v(ws4, f"B{r}", row['Actual'],     bg,     fmt='#,##0', align='right')
#         _v(ws4, f"C{r}", avg_budget,        bg,     fmt='#,##0', align='right')
#         _v(ws4, f"D{r}", var,               bg_var, fmt='#,##0', align='right')
#         _v(ws4, f"E{r}", f"{var_pct:+.1f}%",bg_var, align='center')
#         _v(ws4, f"F{r}", status,            bg_var, align='center')
#         _rh(ws4, r, 20)

#     _border(ws4, f"A2:F{2+len(actual_m)}")
#     for col, w in [(1,14),(2,18),(3,20),(4,18),(5,14),(6,18)]:
#         _cw(ws4, col, w)

#     ws5 = wb.create_sheet("Saudi Events Impact")
#     ws5.sheet_view.showGridLines = False
#     ws5.merge_cells("A1:E1")
#     _h(ws5, "A1", "SAUDI EVENTS — FORECAST IMPACT CALENDAR", DARK, "FFFFFF", True, 12)
#     _rh(ws5, 1, 32)

#     for ci, h in enumerate(["Event","Approx Date","Duration",
#                              "Expected Revenue Impact","CFO Recommendation"], 1):
#         _h(ws5, f"{get_column_letter(ci)}2", h, MID, "FFFFFF")

#     year = datetime.now().year + 1
#     events = [
#         ("Founding Day",       f"22 Feb {year}", "2 days",  "+10% to +15%", "Increase budget allocation"),
#         ("Ramzan",             f"~Mar {year}",   "30 days", "+20% to +35%", "Peak season — plan max capacity"),
#         ("Eid ul Fitr",        f"~Apr {year}",   "4 days",  "+30% to +50%", "Highest revenue — all hands on deck"),
#         ("Hajj Season",        f"~Jun {year}",   "10 days", "+15% to +25%", "Corporate memberships spike"),
#         ("Eid ul Adha",        f"~Jun {year}",   "4 days",  "+25% to +40%", "Second peak — prepare in advance"),
#         ("Muharram",           f"~Jul {year}",   "3 days",  "-5% to +5%",   "Neutral — monitor closely"),
#         ("Saudi National Day", f"23 Sep {year}", "2 days",  "+10% to +20%", "Promotional opportunity"),
#     ]
#     ecols = [LGREEN, LORG, LORG, LBLUE, LORG, GRAY, LBLUE]
#     for ri, (ev, dt, dur, impact, rec) in enumerate(events):
#         r = 3 + ri
#         bg = ecols[ri]
#         _v(ws5, f"A{r}", ev,     bg, bold=True, align='left')
#         _v(ws5, f"B{r}", dt,     bg, align='center')
#         _v(ws5, f"C{r}", dur,    bg, align='center')
#         _v(ws5, f"D{r}", impact, bg, align='center')
#         _v(ws5, f"E{r}", rec,    bg, align='left')
#         _rh(ws5, r, 22)

#     _border(ws5, f"A2:E{2+len(events)}")
#     for col, w in [(1,22),(2,20),(3,14),(4,24),(5,36)]:
#         _cw(ws5, col, w)

#     ws6 = wb.create_sheet("Scenario Planning")
#     ws6.sheet_view.showGridLines = False
#     ws6.merge_cells("A1:D1")
#     _h(ws6, "A1", "SCENARIO PLANNING — BEST / BASE / WORST CASE",
#        DARK, "FFFFFF", True, 12)
#     _rh(ws6, 1, 32)

#     scenarios = [
#         ("BEST CASE",  "All events drive peak, full capacity, no disruptions",
#          fc12['High_Case'].sum(), LGREEN),
#         ("BASE CASE",  "Normal ops with Saudi seasonal adjustments",
#          fc12['Forecast'].sum(),  LBLUE),
#         ("WORST CASE", "Economic slowdown, low event impact",
#          fc12['Low_Case'].sum(),  LRED),
#     ]
#     for ci, h in enumerate(["Scenario","Assumptions","12M Total (SAR)","Avg Monthly (SAR)"],1):
#         _h(ws6, f"{get_column_letter(ci)}2", h, MID, "FFFFFF")

#     for ri, (name, assum, total, bg) in enumerate(scenarios):
#         r = 3 + ri
#         _v(ws6, f"A{r}", name,        bg, bold=True, align='center')
#         _v(ws6, f"B{r}", assum,       bg, align='left')
#         _v(ws6, f"C{r}", total,       bg, fmt='#,##0', bold=True, align='right')
#         _v(ws6, f"D{r}", total/12,    bg, fmt='#,##0', align='right')
#         _rh(ws6, r, 26)

#     ws6.merge_cells("A7:D7")
#     _h(ws6, "A7", "MONTHLY SCENARIO BREAKDOWN", MID, "FFFFFF", True, 11)
#     _rh(ws6, 7, 26)

#     for ci, h in enumerate(["Month","Worst Case (SAR)","Base Case (SAR)","Best Case (SAR)"],1):
#         _h(ws6, f"{get_column_letter(ci)}8", h, DARK, "FFFFFF")

#     for ri, row in fc12.iterrows():
#         r = 9 + ri
#         bg = GRAY if ri % 2 == 0 else WHITE
#         _v(ws6, f"A{r}", str(row['Month']), bg,     align='center')
#         _v(ws6, f"B{r}", row['Low_Case'],   LRED,   fmt='#,##0', align='right')
#         _v(ws6, f"C{r}", row['Forecast'],   LBLUE,  fmt='#,##0', bold=True, align='right')
#         _v(ws6, f"D{r}", row['High_Case'],  LGREEN, fmt='#,##0', align='right')
#         _rh(ws6, r, 20)

#     _border(ws6, f"A8:D{8+len(fc12)}")
#     for col, w in [(1,14),(2,22),(3,22),(4,22)]:
#         _cw(ws6, col, w)

#     buf = io.BytesIO()
#     wb.save(buf)
#     buf.seek(0)
#     return buf


# def main():
#     st.markdown("""
#     <div class="hero">
#         <h1>Mawgif <span class="accent">AI</span> Budget Forecasting</h1>
#         <p>Upload any CSV → AI detects columns → Download Excel forecast instantly</p>
#     </div>
#     """, unsafe_allow_html=True)

#     col1, col2, col3 = st.columns(3)
#     with col1:
#         st.markdown("""<div class="step-card">
#             <div class="step-num">1</div>
#             <h3>Upload CSV</h3>
#             <p>Any format — sales, revenue, membership, amortization data</p>
#         </div>""", unsafe_allow_html=True)
#     with col2:
#         st.markdown("""<div class="step-card">
#             <div class="step-num">2</div>
#             <h3>AI Detects & Forecasts</h3>
#             <p>Auto-detects date, amount, location columns. Trains Prophet model with events</p>
#         </div>""", unsafe_allow_html=True)
#     with col3:
#         st.markdown("""<div class="step-card">
#             <div class="step-num">3</div>
#             <h3>Download Excel</h3>
#             <p>6-sheet professional report: forecast, scenarios, variance, location breakdown</p>
#         </div>""", unsafe_allow_html=True)

#     st.markdown("<hr>", unsafe_allow_html=True)

#     st.markdown("###  Upload Your Data File")
#     uploaded = st.file_uploader("Drop CSV file here", type=['csv'],
#                                  label_visibility='collapsed')

#     if uploaded is None:
#         st.markdown("""<div class="info-box">
#              &nbsp; Upload any CSV with revenue/sales data.
#             AI will automatically detect date, amount, and category columns.
#         </div>""", unsafe_allow_html=True)
#         return

#     try:
#         df = pd.read_csv(uploaded)
#     except Exception as e:
#         st.error(f"Could not read file: {e}")
#         return

#     st.success(f" File loaded — {len(df):,} rows × {len(df.columns)} columns")

#     detected = detect_columns(df)

#     st.markdown("###  AI Column Detection")
#     c1, c2, c3, c4 = st.columns(4)

#     def show_col(container, role, icon, col_name, df):
#         with container:
#             if col_name:
#                 sample = str(df[col_name].dropna().iloc[0]) if len(df) > 0 else "—"
#                 st.markdown(f"""<div class="col-card">
#                     <div class="col-tag">{icon} {role}</div>
#                     <div class="col-name">{col_name}</div>
#                     <div class="col-sample">Sample: {sample[:30]}</div>
#                 </div>""", unsafe_allow_html=True)
#             else:
#                 st.markdown(f"""<div class="col-card">
#                     <div class="col-tag">❓ {role}</div>
#                     <div class="col-name" style="color:#6b7fa3">Not detected</div>
#                     <div class="col-sample">Select manually below</div>
#                 </div>""", unsafe_allow_html=True)

#     show_col(c1, "Date",     "📅", detected['date'],     df)
#     show_col(c2, "Amount",   "💰", detected['amount'],   df)
#     show_col(c3, "Location", "📍", detected['location'], df)
#     show_col(c4, "Category", "🏷️", detected['category'], df)

#     with st.expander("⚙️ Override column detection (optional)"):
#         all_cols = list(df.columns)
#         detected['date']     = st.selectbox("Date column",     all_cols, index=all_cols.index(detected['date'])     if detected['date']     in all_cols else 0)
#         detected['amount']   = st.selectbox("Amount column",   all_cols, index=all_cols.index(detected['amount'])   if detected['amount']   in all_cols else 0)
#         detected['location'] = st.selectbox("Location column (optional)", ['None'] + all_cols,
#                                             index=all_cols.index(detected['location'])+1 if detected['location'] in all_cols else 0)
#         if detected['location'] == 'None':
#             detected['location'] = None

#     if not detected['date'] or not detected['amount']:
#         st.markdown("""<div class="warning-box">
#             ⚠️ Could not detect date or amount column automatically.
#             Please select them manually above.
#         </div>""", unsafe_allow_html=True)
#         return

#     st.markdown("### ⚙️ Forecast Settings")
#     s1, s2 = st.columns(2)
#     with s1:
#         forecast_months = st.slider("Forecast period (months)", 3, 24, 12)
#     with s2:
#         filter_col = detected.get('location') or detected.get('category')
#         filter_val = "All"
#         if filter_col:
#             options = ["All"] + sorted(df[filter_col].dropna().unique().tolist())
#             filter_val = st.selectbox(f"Filter by {filter_col}", options)

#     df_clean = df.copy()
#     df_clean[detected['date']] = pd.to_datetime(df_clean[detected['date']], errors='coerce')
#     df_clean = df_clean.dropna(subset=[detected['date'], detected['amount']])

#     if filter_col and filter_val != "All":
#         df_clean = df_clean[df_clean[filter_col] == filter_val]

#     with st.expander(" Preview data"):
#         st.dataframe(df_clean.head(10), use_container_width=True)

#     st.markdown("---")
#     if st.button(" Generate Budget Forecast"):
#         with st.spinner("AI is analyzing your data and building forecast..."):
#             try:
#                 prophet_df = df_clean.groupby(detected['date'])[detected['amount']].sum().reset_index()
#                 prophet_df.columns = ['ds', 'y']
#                 prophet_df = prophet_df[prophet_df['y'] >= 0].sort_values('ds')

#                 if len(prophet_df) < 30:
#                     st.error("Not enough data — need at least 30 days of history for forecasting.")
#                     return

#                 years = list(range(prophet_df['ds'].dt.year.min(), datetime.now().year + 3))
#                 saudi_holidays = get_saudi_events(years)

#                 model = Prophet(
#                     holidays=saudi_holidays,
#                     yearly_seasonality=True,
#                     weekly_seasonality=True,
#                     daily_seasonality=False,
#                     seasonality_mode='multiplicative',
#                     holidays_prior_scale=15,
#                     changepoint_prior_scale=0.1
#                 )
#                 model.fit(prophet_df)

#                 future   = model.make_future_dataframe(periods=forecast_months * 31)
#                 forecast = model.predict(future)
#                 last_date = prophet_df['ds'].max()
#                 fc_future = forecast[forecast['ds'] > last_date][
#                     ['ds','yhat','yhat_lower','yhat_upper']].copy()

#                 fc_future['month'] = fc_future['ds'].dt.to_period('M')
#                 monthly_fc = fc_future.groupby('month').agg(
#                     {'yhat':'sum','yhat_lower':'sum','yhat_upper':'sum'}).reset_index()
#                 monthly_fc.columns = ['Month','Forecast','Low_Case','High_Case']
#                 monthly_fc['Forecast']  = monthly_fc['Forecast'].clip(lower=0)
#                 monthly_fc['Low_Case']  = monthly_fc['Low_Case'].clip(lower=0)
#                 monthly_fc['High_Case'] = monthly_fc['High_Case'].clip(lower=0)
#                 monthly_fc['Month_str'] = monthly_fc['Month'].astype(str)

#                 st.markdown("""<div class="success-banner">
#                     <h2> Forecast Generated!</h2>
#                     <p>Your AI-powered budget forecast is ready to download</p>
#                 </div>""", unsafe_allow_html=True)

#                 fc12 = monthly_fc.head(forecast_months)
#                 m1, m2, m3, m4 = st.columns(4)
#                 metrics = [
#                     (m1, "Base Forecast (12M)", f"SAR {fc12['Forecast'].sum():,.0f}", ""),
#                     (m2, "High Case (12M)",     f"SAR {fc12['High_Case'].sum():,.0f}", ""),
#                     (m3, "Avg Monthly",         f"SAR {fc12['Forecast'].mean():,.0f}", ""),
#                     (m4, "Peak Month",          str(fc12.loc[fc12['Forecast'].idxmax(),'Month']), ""),
#                 ]
#                 for col, label, val, icon in metrics:
#                     with col:
#                         st.markdown(f"""<div class="metric-box">
#                             <div class="metric-label">{label}</div>
#                             <div class="metric-value">{icon} {val}</div>
#                         </div>""", unsafe_allow_html=True)

#                 st.markdown("<br>", unsafe_allow_html=True)

#                 excel_buf = build_excel(monthly_fc, df_clean, detected, forecast_months)
#                 fname = f"Mawgif_Budget_Forecast_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

#                 st.download_button(
#                     label="⬇️ Download Budget Forecast Excel",
#                     data=excel_buf,
#                     file_name=fname,
#                     mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#                     use_container_width=True
#                 )

#             except Exception as e:
#                 st.error(f"Error: {e}")
#                 st.code(traceback.format_exc())

# if __name__ == "__main__":
#     main()




import streamlit as st
import pandas as pd
import numpy as np
from prophet import Prophet
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings("ignore")
from datetime import datetime, date
import io
import traceback

# ============================================================
# PAGE CONFIG
# ============================================================
st.set_page_config(
    page_title="AI Budget Forecasting",
    page_icon="",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

* { font-family: 'Inter', sans-serif; margin: 0; padding: 0; box-sizing: border-box; }

.stApp { background: #0f1117; color: #e2e8f0; }

/* Sidebar */
div[data-testid="stSidebar"] {
    background: #161b27;
    border-right: 1px solid #1e2d40;
}
div[data-testid="stSidebar"] label { color: #94a3b8 !important; font-size: 0.82rem !important; }

/* Header */
.app-header {
    border-bottom: 1px solid #1e2d40;
    padding: 28px 0 24px 0;
    margin-bottom: 32px;
}
.app-header h1 {
    font-size: 1.6rem;
    font-weight: 600;
    color: #f1f5f9;
    letter-spacing: -0.3px;
}
.app-header p {
    color: #64748b;
    font-size: 0.88rem;
    margin-top: 4px;
}

/* Section label */
.section-label {
    font-size: 0.7rem;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 1px;
    color: #475569;
    margin: 28px 0 12px 0;
}

/* Cards */
.card {
    background: #161b27;
    border: 1px solid #1e2d40;
    border-radius: 8px;
    padding: 20px;
    margin-bottom: 12px;
}
.card-title {
    font-size: 0.78rem;
    color: #64748b;
    font-weight: 500;
    margin-bottom: 6px;
    text-transform: uppercase;
    letter-spacing: 0.5px;
}
.card-value {
    font-size: 1.4rem;
    font-weight: 600;
    color: #f1f5f9;
}
.card-sub {
    font-size: 0.78rem;
    color: #475569;
    margin-top: 3px;
}

/* Detected col pill */
.col-pill {
    background: #1e2d40;
    border: 1px solid #2d3f55;
    border-radius: 6px;
    padding: 12px 16px;
    margin-bottom: 8px;
}
.col-pill-role {
    font-size: 0.7rem;
    color: #3b82f6;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 0.5px;
    margin-bottom: 4px;
}
.col-pill-name { font-size: 0.95rem; color: #e2e8f0; font-weight: 500; }
.col-pill-sample { font-size: 0.78rem; color: #475569; margin-top: 2px; }
.col-pill-miss { font-size: 0.82rem; color: #ef4444; }

/* Alert boxes */
.alert-info {
    background: #1e2d40;
    border-left: 3px solid #3b82f6;
    border-radius: 4px;
    padding: 12px 16px;
    color: #94a3b8;
    font-size: 0.85rem;
    margin: 12px 0;
}
.alert-warn {
    background: #1c1a10;
    border-left: 3px solid #f59e0b;
    border-radius: 4px;
    padding: 12px 16px;
    color: #fbbf24;
    font-size: 0.85rem;
    margin: 12px 0;
}
.alert-success {
    background: #0d1f14;
    border-left: 3px solid #22c55e;
    border-radius: 4px;
    padding: 12px 16px;
    color: #86efac;
    font-size: 0.85rem;
    margin: 12px 0;
}

/* Table */
.forecast-table {
    width: 100%;
    border-collapse: collapse;
    font-size: 0.85rem;
    margin-top: 8px;
}
.forecast-table th {
    background: #1e2d40;
    color: #94a3b8;
    font-weight: 500;
    padding: 10px 14px;
    text-align: left;
    font-size: 0.75rem;
    text-transform: uppercase;
    letter-spacing: 0.5px;
    border-bottom: 1px solid #2d3f55;
}
.forecast-table td {
    padding: 10px 14px;
    border-bottom: 1px solid #1e2d40;
    color: #cbd5e1;
}
.forecast-table tr:hover td { background: #161b27; }
.forecast-table .bold { color: #f1f5f9; font-weight: 600; }
.forecast-table .event-tag {
    background: #1c2e44;
    color: #60a5fa;
    padding: 2px 8px;
    border-radius: 4px;
    font-size: 0.72rem;
    font-weight: 500;
}
.forecast-table .peak { color: #22c55e; font-weight: 600; }
.forecast-table .low  { color: #ef4444; }

/* Divider */
.divider { border: none; border-top: 1px solid #1e2d40; margin: 24px 0; }

/* Button override */
.stButton > button {
    background: #2563eb;
    color: white;
    border: none;
    border-radius: 6px;
    padding: 10px 24px;
    font-size: 0.88rem;
    font-weight: 500;
    width: 100%;
    letter-spacing: 0.2px;
    transition: background 0.15s;
}
.stButton > button:hover { background: #1d4ed8; }

/* Download button */
.stDownloadButton > button {
    background: #16a34a;
    color: white;
    border: none;
    border-radius: 6px;
    padding: 10px 24px;
    font-size: 0.88rem;
    font-weight: 500;
    width: 100%;
}
.stDownloadButton > button:hover { background: #15803d; }

/* Inputs */
.stSelectbox > div, .stTextInput > div > div { 
    background: #1e2d40 !important; 
    border-color: #2d3f55 !important;
    color: #e2e8f0 !important;
}
.stSlider > div { color: #94a3b8; }
label { color: #94a3b8 !important; font-size: 0.82rem !important; }
</style>
""", unsafe_allow_html=True)


# ============================================================
# REGION EVENT DATES — only dates, NO hard-coded impact
# Impact is learned from historical data by Prophet
# ============================================================
REGION_EVENT_DATES = {
    "Saudi Arabia": {
        "Ramzan":             {"month": 3,  "day": 1,  "duration": 30},
        "Eid ul Fitr":        {"month": 4,  "day": 1,  "duration": 4},
        "Eid ul Adha":        {"month": 6,  "day": 16, "duration": 4},
        "Hajj Season":        {"month": 6,  "day": 10, "duration": 10},
        "Saudi National Day": {"month": 9,  "day": 23, "duration": 2},
        "Founding Day":       {"month": 2,  "day": 22, "duration": 2},
        "Muharram":           {"month": 7,  "day": 7,  "duration": 3},
    },
    "Pakistan": {
        "Ramzan":             {"month": 3,  "day": 1,  "duration": 30},
        "Eid ul Fitr":        {"month": 4,  "day": 1,  "duration": 4},
        "Eid ul Adha":        {"month": 6,  "day": 16, "duration": 4},
        "Independence Day":   {"month": 8,  "day": 14, "duration": 1},
        "Christmas / NY":     {"month": 12, "day": 25, "duration": 7},
    },
    "UAE": {
        "Ramzan":               {"month": 3,  "day": 1,  "duration": 30},
        "Eid ul Fitr":          {"month": 4,  "day": 1,  "duration": 4},
        "Eid ul Adha":          {"month": 6,  "day": 16, "duration": 4},
        "UAE National Day":     {"month": 12, "day": 2,  "duration": 2},
        "New Year":             {"month": 1,  "day": 1,  "duration": 2},
        "Dubai Shopping Fest":  {"month": 1,  "day": 15, "duration": 30},
    },
    "Global": {
        "New Year":       {"month": 1,  "day": 1,  "duration": 2},
        "Christmas":      {"month": 12, "day": 25, "duration": 7},
        "Mid-Year":       {"month": 6,  "day": 30, "duration": 1},
    },
}


# ============================================================
# SMART COLUMN DETECTOR
# ============================================================
def detect_columns(df):
    det = {"date": None, "amount": None, "category": None, "location": None}

    date_kw   = ["date", "dt", "time", "period", "month", "year", "day"]
    amount_kw = ["gross_amount", "amount", "revenue", "sales", "net", "value",
                 "income", "payment", "fee", "price", "total", "sum", "amortization"]
    cat_kw    = ["category", "type", "product", "family", "segment",
                 "membership", "plan", "tier"]
    loc_kw    = ["location", "city", "branch", "store", "region", "area", "zone"]

    for col in df.columns:
        cl = col.lower().replace("_", " ")
        if not det["date"] and any(k in cl for k in date_kw):
            try:
                pd.to_datetime(df[col].dropna().iloc[0])
                det["date"] = col
                continue
            except: pass
        if not det["amount"] and any(k in cl for k in amount_kw):
            if pd.api.types.is_numeric_dtype(df[col]):
                det["amount"] = col
                continue
        if not det["category"] and any(k in cl for k in cat_kw):
            if df[col].dtype == object:
                det["category"] = col
                continue
        if not det["location"] and any(k in cl for k in loc_kw):
            if df[col].dtype == object:
                det["location"] = col
                continue

    if not det["date"]:
        for col in df.select_dtypes(include="object").columns:
            try:
                pd.to_datetime(df[col].dropna().iloc[0])
                det["date"] = col
                break
            except: pass

    if not det["amount"]:
        nums = df.select_dtypes(include="number").columns.tolist()
        if nums:
            det["amount"] = nums[-1]

    return det


# ============================================================
# BUILD HOLIDAYS DATAFRAME FOR PROPHET
# ============================================================
def build_holidays(region_events, custom_events, years):
    rows = []

    for name, ev in region_events.items():
        for yr in years:
            try:
                start = pd.Timestamp(f"{yr}-{ev['month']:02d}-{ev['day']:02d}")
                for d in pd.date_range(start=start, periods=ev["duration"], freq="D"):
                    rows.append({
                        "holiday": name, "ds": d,
                        "lower_window": -1, "upper_window": 2
                    })
            except: pass

    for ev in custom_events:
        try:
            start = pd.Timestamp(ev["date"])
            for d in pd.date_range(start=start, periods=ev["duration"], freq="D"):
                rows.append({
                    "holiday": ev["name"], "ds": d,
                    "lower_window": -1, "upper_window": 2
                })
        except: pass

    if rows:
        return pd.DataFrame(rows)
    return pd.DataFrame(columns=["holiday", "ds", "lower_window", "upper_window"])


# ============================================================
# COMPUTE AI-LEARNED EVENT IMPACT FROM HISTORICAL DATA
# ============================================================
def compute_event_impact(prophet_df, holidays_df):
    """
    Let Prophet tell us what it learned about each event.
    No hard-coded values — pure data-driven impact.
    """
    if holidays_df.empty:
        return {}

    mini_model = Prophet(
        holidays=holidays_df,
        yearly_seasonality=True,
        weekly_seasonality=True,
        daily_seasonality=False,
        seasonality_mode="multiplicative",
    )
    mini_model.fit(prophet_df)

    impact = {}
    if hasattr(mini_model, "params") and "holidays" in mini_model.params:
        holiday_params = mini_model.params["holidays"]
        if hasattr(holiday_params, "mean"):
            for i, name in enumerate(holidays_df["holiday"].unique()):
                try:
                    val = float(np.mean(holiday_params[:, i])) * 100
                    impact[name] = round(val, 1)
                except: pass
    return impact


# ============================================================
# EXCEL BUILDER — Clean, professional, 2 sheets only
# ============================================================
def build_excel(monthly_fc, df_clean, det, region_name,
                region_events, custom_events, forecast_years,
                currency, rolling_actuals, learned_impact):

    wb = Workbook()
    wb.remove(wb.active)

    fc = monthly_fc
    total_hist = df_clean[det["amount"]].sum()

    # Month → event name map
    month_event_map = {}
    for name, ev in region_events.items():
        month_event_map[ev["month"]] = name
    for ev in custom_events:
        try:
            m = pd.Timestamp(ev["date"]).month
            month_event_map[m] = ev["name"]
        except: pass

    # Rolling actuals map
    actual_map = {}
    if rolling_actuals:
        for row in rolling_actuals:
            if row["month"] and row["actual"] > 0:
                actual_map[row["month"]] = row["actual"]

    # ── COLORS ──────────────────────────────────────────────
    C_HEADER_BG  = "0f1923"
    C_HEADER_FG  = "94a3b8"
    C_TITLE_BG   = "0a0f1a"
    C_TITLE_FG   = "f1f5f9"
    C_ROW_EVEN   = "111827"
    C_ROW_ODD    = "0f1117"
    C_ACCENT     = "1e3a5f"
    C_GREEN_BG   = "0d2318"
    C_GREEN_FG   = "22c55e"
    C_RED_BG     = "1f0d0d"
    C_RED_FG     = "ef4444"
    C_BLUE_BG    = "0d1f38"
    C_BLUE_FG    = "60a5fa"
    C_WHITE      = "f1f5f9"
    C_MUTED      = "475569"

    def cell(ws, ref, val, bg=None, fg=C_WHITE, bold=False,
             fmt=None, align="left", sz=10, wrap=False):
        c = ws[ref]
        c.value = val
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        c.font = Font(bold=bold, color=fg, size=sz, name="Calibri")
        c.alignment = Alignment(
            horizontal=align, vertical="center", wrap_text=wrap)
        if fmt:
            c.number_format = fmt

    def border_range(ws, rng, color="1e2d40"):
        s = Side(style="thin", color=color)
        for row in ws[rng]:
            for c in row:
                c.border = Border(left=s, right=s, top=s, bottom=s)

    def cw(ws, col, w):
        ws.column_dimensions[get_column_letter(col)].width = w

    def rh(ws, row, h):
        ws.row_dimensions[row].height = h

    # ===========================================================
    # SHEET 1 — BUDGET FORECAST
    # ===========================================================
    ws1 = wb.create_sheet("Budget Forecast")
    ws1.sheet_view.showGridLines = False
    ws1.sheet_properties.tabColor = "2563eb"

    # Title block
    ws1.merge_cells("A1:I1")
    cell(ws1, "A1", "AI Budget Forecasting Framework",
         bg=C_TITLE_BG, fg=C_TITLE_FG, bold=True, sz=13, align="left")
    rh(ws1, 1, 36)

    ws1.merge_cells("A2:I2")
    cell(ws1, "A2",
         f"Generated: {datetime.now().strftime('%d %b %Y')}   |   "
         f"Region: {region_name}   |   "
         f"Forecast: {forecast_years} Year(s)   |   "
         f"Currency: {currency}",
         bg=C_TITLE_BG, fg=C_MUTED, sz=9, align="left")
    rh(ws1, 2, 20)

    # KPI row
    rh(ws1, 3, 10)  # spacer

    kpi_data = [
        ("Historical Revenue",    f"{currency} {total_hist:,.0f}",        "A4", "C4"),
        ("12M Base Forecast",     f"{currency} {fc['base'].sum():,.0f}",   "D4", "F4"),
        ("12M High Case",         f"{currency} {fc['high'].sum():,.0f}",   "G4", "I4"),
    ]
    for label, val, s, e in kpi_data:
        ws1.merge_cells(f"{s}:{e}")
        cell(ws1, s, label, bg=C_ACCENT, fg=C_HEADER_FG, sz=8, align="center")
        rh(ws1, 4, 18)
        r2 = s[0] + "5"
        ws1.merge_cells(f"{r2}:{e[0]}5")
        cell(ws1, r2, val, bg=C_HEADER_BG, fg=C_WHITE, bold=True, sz=11, align="center")
        rh(ws1, 5, 26)

    rh(ws1, 6, 10)  # spacer

    # Column headers
    headers = [
        "Month", "Low Case", "Base Forecast",
        "High Case", "Event", "AI-Learned Impact",
        "MoM Change", "Actual (if entered)", "Variance"
    ]
    for ci, h in enumerate(headers, 1):
        cell(ws1, f"{get_column_letter(ci)}7", h,
             bg=C_HEADER_BG, fg=C_HEADER_FG,
             bold=True, sz=9, align="center")
    rh(ws1, 7, 22)
    border_range(ws1, "A7:I7")

    # Data rows
    prev_base = None
    for ri, (_, row) in enumerate(fc.iterrows()):
        r = 8 + ri
        bg = C_ROW_EVEN if ri % 2 == 0 else C_ROW_ODD

        month_str  = str(row["month"])
        base       = row["base"]
        low        = row["low"]
        high       = row["high"]
        month_num  = row["month_num"]

        # MoM change
        mom = ((base - prev_base) / prev_base * 100) if prev_base else 0
        prev_base  = base

        # Event
        ev_name    = month_event_map.get(month_num, "")
        ev_impact  = learned_impact.get(ev_name, None)
        impact_str = (f"{ev_impact:+.1f}% (learned)"
                      if ev_impact is not None and ev_name else "")

        # Actual & variance
        actual     = actual_map.get(month_str, None)
        variance   = (actual - base) if actual else None

        # Row colors
        mom_bg     = C_GREEN_BG if mom >= 0 else C_RED_BG
        mom_fg     = C_GREEN_FG if mom >= 0 else C_RED_FG
        var_bg     = C_GREEN_BG if (variance or 0) >= 0 else C_RED_BG
        var_fg     = C_GREEN_FG if (variance or 0) >= 0 else C_RED_FG

        cell(ws1, f"A{r}", month_str,  bg=bg, fg=C_WHITE, bold=True, align="center")
        cell(ws1, f"B{r}", low,        bg=bg, fg=C_MUTED, fmt="#,##0", align="right")
        cell(ws1, f"C{r}", base,       bg=bg, fg=C_WHITE, bold=True, fmt="#,##0", align="right")
        cell(ws1, f"D{r}", high,       bg=bg, fg=C_MUTED, fmt="#,##0", align="right")
        cell(ws1, f"E{r}", ev_name,    bg=C_BLUE_BG if ev_name else bg,
             fg=C_BLUE_FG if ev_name else C_MUTED, align="center", sz=9)
        cell(ws1, f"F{r}", impact_str, bg=C_BLUE_BG if impact_str else bg,
             fg=C_BLUE_FG if impact_str else C_MUTED, align="center", sz=9)
        cell(ws1, f"G{r}", f"{mom:+.1f}%", bg=mom_bg, fg=mom_fg,
             bold=True, align="center", sz=9)

        if actual:
            cell(ws1, f"H{r}", actual,    bg=C_ACCENT, fg=C_WHITE, fmt="#,##0", align="right")
            cell(ws1, f"I{r}", variance,  bg=var_bg,   fg=var_fg,  fmt="#,##0", align="right", bold=True)
        else:
            cell(ws1, f"H{r}", "Pending", bg=bg, fg=C_MUTED, align="center", sz=9)
            cell(ws1, f"I{r}", "",        bg=bg)

        rh(ws1, r, 20)

    border_range(ws1, f"A7:I{7 + len(fc)}")

    # Scenario summary at bottom
    summary_row = 9 + len(fc)
    ws1.merge_cells(f"A{summary_row}:I{summary_row}")
    cell(ws1, f"A{summary_row}", "SCENARIO SUMMARY",
         bg=C_HEADER_BG, fg=C_HEADER_FG, bold=True, sz=9, align="center")
    rh(ws1, summary_row, 18)

    sc_row = summary_row + 1
    scenarios = [
        ("Best Case (High)",  fc["high"].sum(),  C_GREEN_BG, C_GREEN_FG, "A", "C"),
        ("Base Forecast",     fc["base"].sum(),  C_BLUE_BG,  C_BLUE_FG,  "D", "F"),
        ("Worst Case (Low)",  fc["low"].sum(),   C_RED_BG,   C_RED_FG,   "G", "I"),
    ]
    for label, val, bg, fg, sc, ec in scenarios:
        ws1.merge_cells(f"{sc}{sc_row}:{ec}{sc_row}")
        cell(ws1, f"{sc}{sc_row}", label,
             bg=bg, fg=fg, sz=8, align="center")
        ws1.merge_cells(f"{sc}{sc_row+1}:{ec}{sc_row+1}")
        cell(ws1, f"{sc}{sc_row+1}",
             f"{currency} {val:,.0f}",
             bg=bg, fg=fg, bold=True, sz=11, align="center")
        rh(ws1, sc_row, 18)
        rh(ws1, sc_row + 1, 26)

    # Column widths
    for c, w in [(1,14),(2,16),(3,18),(4,16),(5,18),(6,20),(7,12),(8,18),(9,16)]:
        cw(ws1, c, w)

    # ===========================================================
    # SHEET 2 — VARIANCE & ACTUALS
    # ===========================================================
    ws2 = wb.create_sheet("Variance Analysis")
    ws2.sheet_view.showGridLines = False
    ws2.sheet_properties.tabColor = "16a34a"

    ws2.merge_cells("A1:G1")
    cell(ws2, "A1", "Variance Analysis — Budget vs Actual",
         bg=C_TITLE_BG, fg=C_TITLE_FG, bold=True, sz=12, align="left")
    rh(ws2, 1, 32)

    ws2.merge_cells("A2:G2")
    cell(ws2, "A2",
         "Historical actual vs average budget target. "
         "AI forecast variance shown where actuals have been entered.",
         bg=C_TITLE_BG, fg=C_MUTED, sz=9, align="left")
    rh(ws2, 2, 18)

    rh(ws2, 3, 8)

    # Historical actual variance
    df2 = df_clean.copy()
    df2[det["date"]] = pd.to_datetime(df2[det["date"]])
    df2["_m"] = df2[det["date"]].dt.to_period("M")
    am = df2.groupby("_m")[det["amount"]].sum().reset_index()
    am.columns = ["Month", "Actual"]
    am = am.tail(24).reset_index(drop=True)
    avg_budget = am["Actual"].mean()

    for ci, h in enumerate(
        ["Month", "Actual Revenue", "Avg Budget Target",
         "Variance", "Variance %", "Status", "Action Required"], 1):
        cell(ws2, f"{get_column_letter(ci)}4", h,
             bg=C_HEADER_BG, fg=C_HEADER_FG, bold=True, sz=9, align="center")
    rh(ws2, 4, 22)

    for ri, row in am.iterrows():
        r = 5 + ri
        var     = row["Actual"] - avg_budget
        var_pct = (var / avg_budget * 100) if avg_budget else 0
        status  = "Above" if var >= 0 else "Below"
        action  = ("Sustain" if var >= 0
                   else "Investigate" if var_pct < -20
                   else "Review")
        bg      = C_ROW_EVEN if ri % 2 == 0 else C_ROW_ODD
        vbg     = C_GREEN_BG if var >= 0 else C_RED_BG
        vfg     = C_GREEN_FG if var >= 0 else C_RED_FG

        cell(ws2, f"A{r}", str(row["Month"]),  bg=bg,  fg=C_WHITE, align="center", bold=True)
        cell(ws2, f"B{r}", row["Actual"],      bg=bg,  fg=C_WHITE, fmt="#,##0", align="right")
        cell(ws2, f"C{r}", avg_budget,         bg=bg,  fg=C_MUTED, fmt="#,##0", align="right")
        cell(ws2, f"D{r}", var,                bg=vbg, fg=vfg,     fmt="#,##0", align="right", bold=True)
        cell(ws2, f"E{r}", f"{var_pct:+.1f}%", bg=vbg, fg=vfg,    align="center", bold=True)
        cell(ws2, f"F{r}", status,             bg=vbg, fg=vfg,     align="center")
        cell(ws2, f"G{r}", action,             bg=bg,  fg=C_MUTED, align="center", sz=9)
        rh(ws2, r, 20)

    border_range(ws2, f"A4:G{4 + len(am)}")
    for c, w in [(1,14),(2,18),(3,20),(4,18),(5,13),(6,12),(7,16)]:
        cw(ws2, c, w)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ============================================================
# MAIN APP
# ============================================================
def main():

    # Header
    st.markdown("""
    <div class="app-header">
        <h1>AI Budget Forecasting Framework</h1>
        <p>Upload any revenue CSV — AI detects columns, learns patterns, generates forecast</p>
    </div>
    """, unsafe_allow_html=True)

    # ── SIDEBAR ─────────────────────────────────────────────
    with st.sidebar:
        st.markdown("**Configuration**")
        st.markdown("---")

        region = st.selectbox(
            "Region",
            ["Saudi Arabia", "Pakistan", "UAE", "Global"],
            help="Used to determine event dates. Impact is learned from your data."
        )

        currency = st.selectbox("Currency", ["SAR", "PKR", "AED", "USD", "EUR", "GBP"])

        forecast_years = st.slider("Forecast period (years)", 1, 10, 1)
        forecast_months = forecast_years * 12

        st.markdown("---")
        st.markdown("**Factors**")
        use_trend    = st.checkbox("Historical Trend",   value=True)
        use_yearly   = st.checkbox("Yearly Seasonality", value=True)
        use_weekly   = st.checkbox("Weekly Seasonality", value=True)
        use_events   = st.checkbox("Events & Holidays",  value=True)

        st.markdown("---")
        st.markdown("**Events**")
        region_events = dict(REGION_EVENT_DATES.get(region, {}))

        # Let user toggle region events
        selected_region_events = {}
        for ev_name, ev_data in region_events.items():
            if st.checkbox(ev_name, value=True, key=f"ev_{ev_name}"):
                selected_region_events[ev_name] = ev_data

        st.markdown("---")
        st.markdown("**Custom Events**")
        if "custom_events" not in st.session_state:
            st.session_state.custom_events = []

        with st.expander("Add custom event"):
            c_name = st.text_input("Name", placeholder="e.g. Company Anniversary")
            c_date = st.date_input("Start date", value=date.today())
            c_dur  = st.number_input("Duration (days)", 1, 90, 1)
            if st.button("Add"):
                if c_name:
                    st.session_state.custom_events.append({
                        "name": c_name,
                        "date": str(c_date),
                        "duration": int(c_dur)
                    })

        for i, ev in enumerate(st.session_state.custom_events):
            st.markdown(f"<small style='color:#64748b'>{ev['name']} — {ev['date']}</small>",
                        unsafe_allow_html=True)

        if st.session_state.custom_events:
            if st.button("Clear custom events"):
                st.session_state.custom_events = []

    # ── STEP 1: UPLOAD ───────────────────────────────────────
    st.markdown('<div class="section-label">Step 1 — Upload Data</div>',
                unsafe_allow_html=True)

    uploaded = st.file_uploader(
        "Upload CSV file", type=["csv"],
        label_visibility="collapsed"
    )

    if not uploaded:
        st.markdown("""<div class="alert-info">
            Upload any CSV with revenue or sales data.
            AI will automatically detect date, amount, and category columns.
        </div>""", unsafe_allow_html=True)
        return

    try:
        df = pd.read_csv(uploaded)
    except Exception as e:
        st.error(f"Could not read file: {e}")
        return

    st.markdown(f"""<div class="alert-success">
        File loaded — {len(df):,} rows, {len(df.columns)} columns
    </div>""", unsafe_allow_html=True)

    # ── STEP 2: COLUMN DETECTION ─────────────────────────────
    st.markdown('<div class="section-label">Step 2 — Column Detection</div>',
                unsafe_allow_html=True)

    detected = detect_columns(df)

    c1, c2, c3, c4 = st.columns(4)
    for cont, role, key in [
        (c1, "Date Column",     "date"),
        (c2, "Amount Column",   "amount"),
        (c3, "Location Column", "location"),
        (c4, "Category Column", "category"),
    ]:
        with cont:
            col = detected[key]
            if col:
                sample = str(df[col].dropna().iloc[0])[:20] if len(df) > 0 else ""
                st.markdown(f"""<div class="col-pill">
                    <div class="col-pill-role">{role}</div>
                    <div class="col-pill-name">{col}</div>
                    <div class="col-pill-sample">Sample: {sample}</div>
                </div>""", unsafe_allow_html=True)
            else:
                st.markdown(f"""<div class="col-pill">
                    <div class="col-pill-role">{role}</div>
                    <div class="col-pill-miss">Not detected — select below</div>
                </div>""", unsafe_allow_html=True)

    with st.expander("Override column selection"):
        all_cols = list(df.columns)
        detected["date"]   = st.selectbox("Date column",   all_cols,
            index=all_cols.index(detected["date"])   if detected["date"]   in all_cols else 0)
        detected["amount"] = st.selectbox("Amount column", all_cols,
            index=all_cols.index(detected["amount"]) if detected["amount"] in all_cols else 0)
        lo = ["None"] + all_cols
        ls = st.selectbox("Location column", lo,
            index=all_cols.index(detected["location"]) + 1 if detected["location"] in all_cols else 0)
        detected["location"] = None if ls == "None" else ls

    if not detected["date"] or not detected["amount"]:
        st.markdown("""<div class="alert-warn">
            Date and amount columns are required. Please select them above.
        </div>""", unsafe_allow_html=True)
        return

    # ── STEP 3: ROLLING ACTUALS ──────────────────────────────
    st.markdown('<div class="section-label">Step 3 — Rolling Actuals (Optional)</div>',
                unsafe_allow_html=True)

    st.markdown("""<div class="alert-info">
        Enter actual figures as they come in each month.
        Variance will be calculated and shown in the forecast output.
    </div>""", unsafe_allow_html=True)

    rolling_actuals = []
    with st.expander("Enter actual monthly figures"):
        for i in range(6):
            ca, cb = st.columns(2)
            with ca:
                m = st.text_input(f"Month {i+1}", placeholder="e.g. 2026-03", key=f"rm{i}")
            with cb:
                v = st.number_input("Actual amount", min_value=0.0, key=f"rv{i}", step=1000.0)
            if m and v > 0:
                rolling_actuals.append({"month": m, "actual": v})

    with st.expander("Preview data"):
        st.dataframe(df.head(10), use_container_width=True)

    # ── STEP 4: GENERATE ─────────────────────────────────────
    st.markdown('<div class="section-label">Step 4 — Generate Forecast</div>',
                unsafe_allow_html=True)

    col_a, col_b, col_c = st.columns(3)
    with col_a:
        st.markdown(f"""<div class="card">
            <div class="card-title">Region</div>
            <div class="card-value">{region}</div>
        </div>""", unsafe_allow_html=True)
    with col_b:
        st.markdown(f"""<div class="card">
            <div class="card-title">Forecast Period</div>
            <div class="card-value">{forecast_years} Year(s)</div>
            <div class="card-sub">{forecast_months} months</div>
        </div>""", unsafe_allow_html=True)
    with col_c:
        ev_count = len(selected_region_events) + len(st.session_state.custom_events)
        st.markdown(f"""<div class="card">
            <div class="card-title">Events Selected</div>
            <div class="card-value">{ev_count}</div>
            <div class="card-sub">Impact learned from data</div>
        </div>""", unsafe_allow_html=True)

    if st.button("Generate Forecast"):
        with st.spinner("Training AI model on your data..."):
            try:
                # Prepare
                df_c = df.copy()
                df_c[detected["date"]] = pd.to_datetime(df_c[detected["date"]], errors="coerce")
                df_c = df_c.dropna(subset=[detected["date"], detected["amount"]])

                prophet_df = (
                    df_c.groupby(detected["date"])[detected["amount"]]
                    .sum()
                    .reset_index()
                )
                prophet_df.columns = ["ds", "y"]
                prophet_df = prophet_df[prophet_df["y"] >= 0].sort_values("ds")

                if len(prophet_df) < 30:
                    st.error("Need at least 30 days of historical data.")
                    return

                years = list(range(
                    prophet_df["ds"].dt.year.min(),
                    datetime.now().year + forecast_years + 2
                ))

                # Build holidays — dates only, no impact values
                hdf = pd.DataFrame()
                if use_events:
                    hdf = build_holidays(
                        selected_region_events if use_events else {},
                        st.session_state.custom_events if use_events else [],
                        years
                    )

                # Train Prophet — AI learns impact from data
                model = Prophet(
                    holidays=hdf if not hdf.empty else None,
                    yearly_seasonality=use_yearly,
                    weekly_seasonality=use_weekly,
                    daily_seasonality=False,
                    seasonality_mode="multiplicative",
                    changepoint_prior_scale=0.1 if use_trend else 0.01,
                )
                model.fit(prophet_df)

                # Learn event impact from model — no hard coding
                learned_impact = {}
                if use_events and not hdf.empty:
                    try:
                        components = model.predict(
                            model.make_future_dataframe(periods=30))
                        holiday_cols = [
                            c for c in components.columns
                            if c in hdf["holiday"].unique()
                        ]
                        for ev_name in holiday_cols:
                            non_zero = components[components[ev_name] != 0][ev_name]
                            if len(non_zero) > 0:
                                learned_impact[ev_name] = round(
                                    float(non_zero.mean()) * 100, 1)
                    except: pass

                # Forecast
                future   = model.make_future_dataframe(periods=forecast_months * 31)
                forecast = model.predict(future)

                last_date = prophet_df["ds"].max()
                ff = forecast[forecast["ds"] > last_date][
                    ["ds", "yhat", "yhat_lower", "yhat_upper"]
                ].copy()
                ff["month"] = ff["ds"].dt.to_period("M")

                mfc = ff.groupby("month").agg(
                    {"yhat": "sum", "yhat_lower": "sum", "yhat_upper": "sum"}
                ).reset_index()
                mfc.columns = ["month", "base", "low", "high"]
                mfc["base"]     = mfc["base"].clip(lower=0)
                mfc["low"]      = mfc["low"].clip(lower=0)
                mfc["high"]     = mfc["high"].clip(lower=0)
                mfc["month_str"]= mfc["month"].astype(str)
                mfc["month_num"]= mfc["month"].apply(lambda x: x.month)
                mfc = mfc.head(forecast_months)

                # Show results
                st.markdown("""<div class="alert-success">
                    Forecast generated. Download the Excel report below.
                </div>""", unsafe_allow_html=True)

                # Summary cards
                m1, m2, m3, m4 = st.columns(4)
                for cont, label, val, sub in [
                    (m1, "Base Forecast",
                     f"{currency} {mfc['base'].sum():,.0f}",
                     f"{forecast_years}Y total"),
                    (m2, "High Case",
                     f"{currency} {mfc['high'].sum():,.0f}", ""),
                    (m3, "Low Case",
                     f"{currency} {mfc['low'].sum():,.0f}", ""),
                    (m4, "Monthly Average",
                     f"{currency} {mfc['base'].mean():,.0f}", "base case"),
                ]:
                    with cont:
                        st.markdown(f"""<div class="card">
                            <div class="card-title">{label}</div>
                            <div class="card-value">{val}</div>
                            <div class="card-sub">{sub}</div>
                        </div>""", unsafe_allow_html=True)

                # Forecast table preview
                st.markdown('<div class="section-label">Forecast Preview</div>',
                            unsafe_allow_html=True)

                month_event_map = {}
                for name, ev in selected_region_events.items():
                    month_event_map[ev["month"]] = name
                for ev in st.session_state.custom_events:
                    try:
                        month_event_map[pd.Timestamp(ev["date"]).month] = ev["name"]
                    except: pass

                actual_map = {r["month"]: r["actual"] for r in rolling_actuals}

                rows_html = ""
                prev_base = None
                for _, row in mfc.iterrows():
                    base     = row["base"]
                    low      = row["low"]
                    high     = row["high"]
                    mom      = ((base - prev_base) / prev_base * 100) if prev_base else 0
                    prev_base = base
                    ev_name  = month_event_map.get(row["month_num"], "")
                    impact   = learned_impact.get(ev_name)
                    imp_str  = f"{impact:+.1f}%" if impact is not None and ev_name else ""
                    actual   = actual_map.get(row["month_str"])
                    var      = actual - base if actual else None

                    mom_cls = "peak" if mom >= 5 else "low" if mom < -5 else ""
                    rows_html += f"""
                    <tr>
                        <td class="bold">{row['month_str']}</td>
                        <td style="color:#475569">{currency} {low:,.0f}</td>
                        <td class="bold">{currency} {base:,.0f}</td>
                        <td style="color:#475569">{currency} {high:,.0f}</td>
                        <td>{"<span class='event-tag'>" + ev_name + "</span>" if ev_name else ""}</td>
                        <td style="color:#60a5fa;font-size:0.82rem">{imp_str}</td>
                        <td class="{mom_cls}">{mom:+.1f}%</td>
                        <td>{currency + " " + f"{actual:,.0f}" if actual else "<span style='color:#475569'>Pending</span>"}</td>
                        <td {"style='color:#22c55e'" if var and var>=0 else "style='color:#ef4444'" if var else ""}>
                            {currency + " " + f"{var:,.0f}" if var else ""}
                        </td>
                    </tr>"""

                st.markdown(f"""
                <table class="forecast-table">
                    <thead><tr>
                        <th>Month</th><th>Low Case</th><th>Base Forecast</th>
                        <th>High Case</th><th>Event</th><th>AI-Learned Impact</th>
                        <th>MoM Change</th><th>Actual</th><th>Variance</th>
                    </tr></thead>
                    <tbody>{rows_html}</tbody>
                </table>
                """, unsafe_allow_html=True)

                # Download
                st.markdown("<br>", unsafe_allow_html=True)
                excel_buf = build_excel(
                    mfc, df_c, detected, region,
                    selected_region_events,
                    st.session_state.custom_events,
                    forecast_years, currency,
                    rolling_actuals, learned_impact
                )
                fname = f"Budget_Forecast_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
                st.download_button(
                    "Download Excel Report",
                    data=excel_buf,
                    file_name=fname,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

            except Exception as e:
                st.error(f"Error: {e}")
                st.code(traceback.format_exc())


if __name__ == "__main__":
    main()